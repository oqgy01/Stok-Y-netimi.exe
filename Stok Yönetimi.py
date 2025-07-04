#region // Kütüphaneler

import requests
from bs4 import BeautifulSoup
import pandas as pd
from io import BytesIO
import re
from colorama import init, Fore, Style
from datetime import datetime, timedelta
import os
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
import xlsxwriter
import gc
from supabase import create_client, Client
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import time
import http.client
import json
import datetime
import warnings
import copy
import colorama
from copy import copy
from openpyxl.worksheet.table import Table, TableStyleInfo
import sys
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
init(autoreset=True)
warnings.filterwarnings("ignore")
colorama.init(autoreset=True)

#endregion

#region // Entegrasyondan Önce mi Sonra mı Kontrolü


# Supabase için gerekli kütüphane


# Supabase bağlantı bilgileri
SUPABASE_URL = "https://zmvsatlvobhdaxxgtoap.supabase.co"
SUPABASE_KEY = (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9."
    "eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InptdnNhdGx2b2JoZGF4eGd0b2FwIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NDAxNzIxMzksImV4cCI6MjA1NTc0ODEzOX0."
    "lJLudSfixMbEOkJmfv22MsRLofP7ZjFkbGj26xF3dts"
)


def list_detail_with_http_client():
    # 0) Öncelikle Supabase üzerinden kategorisatis tablosundaki en büyük 'date' değerini kontrol ediyoruz
    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
    response = (
        supabase.table("kategorisatis")
        .select("date")
        .order("date", desc=True)
        .limit(1)
        .execute()
    )

    if not response.data:
        print("kategorisatis tablosunda herhangi bir kayıt bulunamadı.")
        return

    max_date_str = response.data[0]["date"]  # Örnek format: "2025-03-25"
    max_date = datetime.datetime.strptime(max_date_str, "%Y-%m-%d").date()
    today = datetime.datetime.now().date()

    # Tarih kontrolü
    if max_date == today:
        print("\033[92mEntegrasyondan Sonraki Listeyi Çekiyorsunuz !\033[0m")
    else:
        print("\033[91mDikkat Entegrasyondan Önceki Listeyi Çekiyorsunuz ! (Entegrasyondan Önceki Listeyi Çekmek Stok Adetlerinin Güncelliğini Önemli Ölçüde Etkiler) \033[0m")

    # Supabase kontrolünden SONRA HTTP işlemlerine devam edebilirsiniz.
    # ----------------------------------------------------------------

    # 1) Giriş (login) isteği
    conn = http.client.HTTPSConnection("task.haydigiy.com")

    login_payload = {
        "apiKey": "MypGcaEInEOTzuYQydgDHQ",
        "secretKey": "jRqliBLDPke76YhL_WL5qg",
        "emailOrPhone": "mustafa_kod@haydigiy.com",
        "password": "123456"
    }
    login_headers = {
        "Content-Type": "application/json",
        "Accept": "application/json"
    }

    conn.request(
        "POST",
        "/api/customer/login",
        body=json.dumps(login_payload),
        headers=login_headers
    )

    res = conn.getresponse()
    data = res.read().decode("utf-8")

    if res.status != 200:
        print("Giriş başarısız:", data)
        return

    login_data = json.loads(data)
    token = login_data.get("data", {}).get("token")
    if not token:
        print("Token alınamadı. Dönen veri:", login_data)
        return

    # 2) list-detail isteği
    conn = http.client.HTTPSConnection("task.haydigiy.com")

    list_detail_payload = {
        "searchTerm": None,
        "inCategoryIds": [],
        "includeInSubCategories": True,
        "notInCategoryIds": [],
        "includeNotInSubCategories": True,
        "inManufacturerIds": [],
        "notInManufacturerIds": [],
        "inVendorIds": [],
        "notInVendorIds": [],
        "inProductTagIds": [],
        "notInProductTagIds": [],
        "published": None,
        "minStock": None,
        "maxStock": None,
        "minPrice": None,
        "maxPrice": None,
        "minProductCost": None,
        "maxProductCost": None,
        "pageIndex": 1,
        "pageSize": 1
    }

    list_detail_headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "Accept": "application/json"
    }

    conn.request(
        "POST",
        "/adminapi/product/list-detail",
        body=json.dumps(list_detail_payload),
        headers=list_detail_headers
    )

    res = conn.getresponse()
    data = res.read().decode("utf-8")

    if res.status != 200:
        print("List-detail isteği başarısız:", data)
        return

    response_data = json.loads(data)
    items = response_data.get("data", [])

    if not items:
        print("Herhangi bir ürün kaydı bulunamadı.")
        return

    # Burada artık 'createdOn' kontrolü yerine Supabase tarihini kullandığımız için
    # ek bir kontrol yapmamıza gerek yok. İsterseniz 'createdOn' alanını da kullanabilirsiniz.



if __name__ == "__main__":
    list_detail_with_http_client()


#endregion





# ------------------------------------------------------------
# Yardımcı Fonksiyonlar
# ------------------------------------------------------------

def parse_decimal_string(value):
    """
    Örneğin "1,0000" veya "99,9900" gibi virgüllü ondalıkları float'a çevirir.
    "129,9900" -> 129.99
    """
    s = str(value).strip()      # string'e çevir, boşlukları temizle
    s = s.replace(',', '.')     # Virgülleri noktaya çeviru
    try:
        return float(s)
    except ValueError:
        return 0.0

def extract_product_code(urun_adi):
    """
    Örnek: "TEAM BRODE - 62532. Kırmızı" -> "62532"
    '- (\\d+)\\.' kalıbını arar; yoksa None döner.
    """
    match = re.search(r' - (\d+)\.', urun_adi)
    return match.group(1) if match else None

def extract_color(urun_adi):
    """
    Örnek: "TEAM BRODE - 62532. Kırmızı" -> "Kırmızı"
    Yöntem: ' - ' bazında böldükten sonra ilk parçanın son kelimesini alıyoruz.
    Ama veriniz farklı yapıdaysa lütfen uyarlayın.
    """
    parts = re.split(r' - ', urun_adi)
    if len(parts) > 0:
        before_part = parts[0].strip()
        words = before_part.split()
        if words:
            return words[-1]
    return None





# ------------------------------------------------------------
# 1) Supabase üzerinden Ürün Listesi (tum_urun_listesi.xlsx) İndirme ve İşleme
# ------------------------------------------------------------
SUPABASE_URL = "https://zmvsatlvobhdaxxgtoap.supabase.co"
SUPABASE_KEY = (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9."
    "eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InptdnNhdGx2b2JoZGF4eGd0b2FwIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NDAxNzIxMzksImV4cCI6MjA1NTc0ODEzOX0."
    "lJLudSfixMbEOkJmfv22MsRLofP7ZjFkbGj26xF3dts"
)

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# "tum_urun_listesi.xlsx" indirilir (belleğe)
response = supabase.storage.from_("tum_urun_listesi").download("tum_urun_listesi.xlsx")
temp_df = pd.read_excel(BytesIO(response))

# Kullanmak istediğimiz sütunlar
selected_columns = [
    "ModelKodu",
    "StokKodu",
    "UrunAdi",
    "StokAdedi",
    "AlisFiyati",
    "SatisFiyati",
    "Kategori",
    "Resim",
    "AramaTerimleri",
    "MorhipoKodu",
    "VaryasyonMorhipoKodu",
    "HepsiBuradaKodu",
    "Marka",
    "N11Kodu",
    "VaryasyonGittiGidiyorKodu",
    "TrendyolKodu",
    "VaryasyonTrendyolKodu",
    "Ozellik",
    "Varyasyon",
    "VaryasyonAmazonKodu",
]
df = temp_df[selected_columns].copy()

# --------------------------------------
# Varyasyon kolonu için özel toplama işlemi
# --------------------------------------
def size_key(v):
    predefined_order = ["XXS", "XS", "S", "M", "L", "XL", "XXL", "XXXL"]
    v = v.strip()
    if v in predefined_order:
        return (0, predefined_order.index(v))
    try:
        return (1, int(v))
    except ValueError:
        return (2, v)

def aggregate_variations(grp):
    grp = grp.copy()
    grp["clean_var"] = grp["Varyasyon"].str.replace("Beden:", "", regex=False).str.strip()
    agg = grp.groupby("clean_var", as_index=False)["StokAdedi"].sum()
    agg = agg.sort_values(by="clean_var", key=lambda col: col.map(lambda x: size_key(x)))
    return " // ".join(f"{row['clean_var']} : {row['StokAdedi']}" for _, row in agg.iterrows())

agg_variations = df.groupby("UrunAdi").apply(aggregate_variations).reset_index(name="Varyasyon_Agg")
df = df.drop(columns=["Varyasyon"]).merge(agg_variations, on="UrunAdi", how="left")
df.rename(columns={"Varyasyon_Agg": "Varyasyon"}, inplace=True)

# --------------------------------------
# Diğer işlemler: Benzersiz kayıtlar ve StokAdedi toplama
# --------------------------------------
unique_cols = [c for c in selected_columns if c not in ["StokAdedi", "Varyasyon"]]
unique_df = df.drop(columns=["StokAdedi"]).drop_duplicates(subset=unique_cols)

stokadedi_sums = df.groupby("UrunAdi")["StokAdedi"].sum().clip(lower=0).reset_index()
final_df = pd.merge(unique_df, stokadedi_sums, on="UrunAdi", how="left")

# --------------------------------------
# VaryasyonAmazonKodu kolonuna link ekleme (.0 son ekini at)
# --------------------------------------
def make_link(pid):
    if pd.isna(pid):
        return ""
    # sayısal değerleri ".0" sız biçime çevir
    try:
        pid_str = str(int(float(pid)))
    except (ValueError, TypeError):
        pid_str = str(pid).split(".")[0]
    return f'=HYPERLINK("https://hgstokyonetim.com/api/sigaraurun?productId={pid_str}", "Sigara Ekle/Çıkar")'

final_df["VaryasyonAmazonKodu"] = final_df["VaryasyonAmazonKodu"].apply(make_link)
































# ------------------------------------------------------------
# 2) Satış Raporu İndirme (Selenium) + Bellekte Parse Edip final_df'ye Ekleme
# ------------------------------------------------------------

colorama.init()

chrome_options = Options()
chrome_options.add_argument("--headless")
chrome_options.add_argument("--incognito")
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")

service = Service()
driver = webdriver.Chrome(service=service, options=chrome_options)

username = "mustafa_kod@haydigiy.com"
password = "123456"
login_url = "https://www.siparis.haydigiy.com/kullanici-giris/?ReturnUrl=%2Fadmin"
desired_page_url = "https://www.siparis.haydigiy.com/admin/exportorder/edit/154/"

try:
    # 2.1) Giriş
    driver.get(login_url)
    time.sleep(2)
    driver.find_element(By.NAME, "EmailOrPhone").send_keys(username)
    driver.find_element(By.NAME, "Password").send_keys(password)
    driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()
    time.sleep(3)

    # 2.2) İlgili sayfaya gidip tarih ayarlarını girmek
    driver.get(desired_page_url)
    time.sleep(2)

    yesterday = datetime.datetime.now() - datetime.timedelta(days=1)
    formatted_date_no_leading = f"{yesterday.day}.{yesterday.month}.{yesterday.year}"

    end_date_input = driver.find_element(By.ID, "EndDate")
    end_date_input.clear()
    end_date_input.send_keys(formatted_date_no_leading)

    start_date_input = driver.find_element(By.ID, "StartDate")
    start_date_input.clear()
    start_date_input.send_keys(formatted_date_no_leading)

    save_button = driver.find_element(By.CSS_SELECTOR, 'button.btn.btn-primary[name="save"]')
    save_button.click()

    time.sleep(10)

except Exception as e:
    print(colorama.Fore.RED + f"Giriş veya tarih ayarı sırasında hata: {e}" + colorama.Style.RESET_ALL)
    raise  # Hata sonrası kodun tamamen durması için hatayı tekrar fırlatıyoruz

finally:
    driver.quit()

# 2.3) Bellekte Satış Raporu verisini alalım
url = "https://www.siparis.haydigiy.com/FaprikaOrderXls/GZPCKE/1/"
resp = requests.get(url)
sales_df = pd.read_excel(BytesIO(resp.content))

# 2.4) "Adet" ve "ToplamFiyat" float'a çevirelim
columns_to_keep = ["UrunAdi", "Adet", "ToplamFiyat"]
sales_df = sales_df[columns_to_keep]
sales_df["Adet"] = sales_df["Adet"].apply(parse_decimal_string)
sales_df["ToplamFiyat"] = sales_df["ToplamFiyat"].apply(parse_decimal_string)

# 2.5) UrunAdi bazında sum
sales_sums = sales_df.groupby("UrunAdi", as_index=False).agg({
    "Adet": "sum",
    "ToplamFiyat": "sum"
})

# 2.6) final_df ile birleştir => final_merged_df
final_merged_df = pd.merge(final_df, sales_sums, on="UrunAdi", how="left")
final_merged_df["Adet"] = final_merged_df["Adet"].fillna(0)
final_merged_df["ToplamFiyat"] = final_merged_df["ToplamFiyat"].fillna(0)

# ------------------------------------------------------------
# 3) "UrunAdi Duzenleme" ve "UrunAdi ve Renk" Kolonlarını Oluşturma
# ------------------------------------------------------------

final_merged_df["UrunAdi Duzenleme"] = final_merged_df["UrunAdi"].apply(extract_product_code)
final_merged_df["UrunAdi Duzenleme"] = final_merged_df["UrunAdi Duzenleme"].fillna("").astype(str)

def build_urun_adi_ve_renk(row):
    """
    "UrunAdi ve Renk" = UrunAdi Duzenleme + " - " + Renk
    """
    product_code_str = row["UrunAdi Duzenleme"]
    color_str = extract_color(row["UrunAdi"])  # Örnek: "Kırmızı"
    if not color_str:
        color_str = ""
    return product_code_str + " - " + color_str

final_merged_df["UrunAdi ve Renk"] = final_merged_df.apply(build_urun_adi_ve_renk, axis=1)

# ------------------------------------------------------------
# YENİ EKLEME: "urunyonetimi" Tablosundan "id, urunkodu, renk, alisfiyati" Kolonlarını Çekip
# ID En Büyük Olan Kaydı Baz Alarak "Satın Alma Fiyatı" Sütununu Oluşturma
# ------------------------------------------------------------

# 1) Supabase'den veriyi alıyoruz
response_data_fiyat = supabase.table("urunyonetimi").select("id, urunkodu, renk, alisfiyati").execute()
df_urunyonetimi_fiyat = pd.DataFrame(response_data_fiyat.data)

# 2) urunkodu ve renk'i birleştireceğimiz kolon
df_urunyonetimi_fiyat["urunkodu_renk"] = df_urunyonetimi_fiyat["urunkodu"].astype(str) + " - " + df_urunyonetimi_fiyat["renk"].astype(str)

# 3) Aynı urunkodu_renk'e sahip birden fazla satır varsa, ID değeri en büyük olanı al
df_urunyonetimi_fiyat.sort_values("id", ascending=False, inplace=True)
df_urunyonetimi_fiyat.drop_duplicates(subset=["urunkodu_renk"], keep="first", inplace=True)

# 4) final_merged_df ile merge
final_merged_df = final_merged_df.merge(
    df_urunyonetimi_fiyat[["urunkodu_renk", "alisfiyati"]],
    how="left",
    left_on="UrunAdi ve Renk",
    right_on="urunkodu_renk"
)

# 5) Gelen "fiyat" kolonunu "Satın Alma Fiyatı" olarak yeniden adlandırıyoruz
final_merged_df.rename(columns={"alisfiyati": "Son Satın Alma Fiyatı"}, inplace=True)

# 6) "urunkodu_renk" geçici kolonunu atıyoruz
final_merged_df.drop("urunkodu_renk", axis=1, inplace=True)

# 7) "Satın Alma Fiyatı" kolonunu "AlisFiyati" kolonunun hemen sağına yerleştiriyoruz
cols = list(final_merged_df.columns)
if "AlisFiyati" in cols and "Son Satın Alma Fiyatı" in cols:
    alis_idx = cols.index("AlisFiyati")
    satinalma_idx = cols.index("Son Satın Alma Fiyatı")
    col_to_move = cols.pop(satinalma_idx)
    cols.insert(alis_idx + 1, col_to_move)
    final_merged_df = final_merged_df[cols]

# ------------------------------------------------------------
# 4) GMT ve SİTA Verilerini Çekme (Supabase tablosu "urunyonetimi")
#    – tedarikci kolonu eklendi
#    – firmakodu çıkarıldı
#    – acilanadet == 0 filtresi eklendi
# ------------------------------------------------------------

all_data = []
start      = 0
page_size  = 1000

while True:
    end = start + page_size - 1
    response_data = (
        supabase.table("urunyonetimi")
        .select("urunkodu, renk, tedarikci, acilanadet, acilmamisadet, gmtsitalabel")
        .in_("gmtsitalabel", ["GMT", "SİTA", "Yarım GMT"])
        .eq("acilanadet", 0)                          # YENİ KOŞUL
        .range(start, end)
        .execute()
    )
    data = response_data.data
    if not data:
        break
    all_data.extend(data)
    start += page_size

df_gmtsita = pd.DataFrame(all_data)

# firmakodu → tedarikci sütunundaki son noktadan sonraki sayı
df_gmtsita["firmakodu"] = (
    df_gmtsita["tedarikci"]
      .astype(str)
      .str.extract(r'\.(\d+)\b', expand=False)
)

# Renk ilk harf büyük
df_gmtsita["renk"] = df_gmtsita["renk"].apply(lambda x: x.capitalize() if isinstance(x, str) else x)

# GMT + Yarım GMT
df_gmt = df_gmtsita[df_gmtsita["gmtsitalabel"].isin(["GMT", "Yarım GMT"])]
df_gmt_grouped = (
    df_gmt.groupby(["urunkodu", "renk", "firmakodu"], as_index=False)["acilmamisadet"].sum()
)

# GMT
col_gmt_urun_kodu = (
    pd.to_numeric(df_gmt_grouped["urunkodu"], errors="coerce")
      .fillna(0)
      .round()
      .astype("Int64")
)

df_gmt_final = pd.DataFrame({
    "GMT Ürün Kodu": col_gmt_urun_kodu,
    "Firma Kodu"   : df_gmt_grouped["firmakodu"],
    "GMT Ürün Adı" : df_gmt_grouped["urunkodu"].astype(str) + " - " +
                     df_gmt_grouped["renk"].astype(str)     + " - " +
                     df_gmt_grouped["firmakodu"].astype(str),
    "GMT Stok Adedi": df_gmt_grouped["acilmamisadet"],
})

# SİTA
df_sita = df_gmtsita[df_gmtsita["gmtsitalabel"] == "SİTA"]
df_sita_grouped = (
    df_sita.groupby(["urunkodu", "renk", "firmakodu"], as_index=False)["acilmamisadet"].sum()
)

col_sita_urun_kodu = (
    pd.to_numeric(df_sita_grouped["urunkodu"], errors="coerce")
      .fillna(0)
      .round()
      .astype("Int64")
)

df_sita_final = pd.DataFrame({
    "SİTA Ürün Kodu": col_sita_urun_kodu,
    "Firma Kodu"    : df_sita_grouped["firmakodu"],
    "SİTA Ürün Adı" : df_sita_grouped["urunkodu"].astype(str) + " - " +
                      df_sita_grouped["renk"].astype(str)     + " - " +
                      df_sita_grouped["firmakodu"].astype(str),
    "SİTA Stok Adedi": df_sita_grouped["acilmamisadet"],
})

# ------------------------------------------------------------
# 5) GMT ve SİTA Verilerini Ana Tabloya Çektirme (Etopla Mantığı)
#    – eşleştirme anahtarına firmakodu eklendi
# ------------------------------------------------------------

df_calisma_alani = final_merged_df.copy()

# final_merged_df tarafında firmakodu sütunu oluştur
df_calisma_alani["FirmaKodu"] = (
    df_calisma_alani["UrunAdi"]
      .astype(str)
      .str.extract(r'\.(\d+)\b', expand=False)
)

# birleşik anahtar
df_calisma_alani["UrunAdi_Renk_Firma"] = (
    df_calisma_alani["UrunAdi ve Renk"] + " - " + df_calisma_alani["FirmaKodu"]
)

used_gmt_indices_step1  = []
used_sita_indices_step1 = []

# 5.1) Ürün adı + renk + firma kodu ile direkt eşleştirme
for idx, row in df_calisma_alani.iterrows():
    key = row["UrunAdi_Renk_Firma"]

    # GMT
    match_gmt = df_gmt_final[df_gmt_final["GMT Ürün Adı"] == key]
    if not match_gmt.empty:
        m_idx = match_gmt.index[0]
        df_calisma_alani.at[idx, "GMT Stok Adedi"] = match_gmt.iloc[0]["GMT Stok Adedi"]
        used_gmt_indices_step1.append(m_idx)
    else:
        df_calisma_alani.at[idx, "GMT Stok Adedi"] = None

    # SİTA
    match_sita = df_sita_final[df_sita_final["SİTA Ürün Adı"] == key]
    if not match_sita.empty:
        m_idx = match_sita.index[0]
        df_calisma_alani.at[idx, "SİTA Stok Adedi"] = match_sita.iloc[0]["SİTA Stok Adedi"]
        used_sita_indices_step1.append(m_idx)
    else:
        df_calisma_alani.at[idx, "SİTA Stok Adedi"] = None

df_gmt_final  = df_gmt_final.drop(used_gmt_indices_step1).reset_index(drop=True)
df_sita_final = df_sita_final.drop(used_sita_indices_step1).reset_index(drop=True)

used_gmt_indices_step2  = []
used_sita_indices_step2 = []

# 5.2) Kod + firma kodu ile eşleştirme
for idx, row in df_calisma_alani.iterrows():
    urunkodu   = row["UrunAdi Duzenleme"]
    firma_kodu = row["FirmaKodu"]

    # GMT
    if pd.isna(row["GMT Stok Adedi"]) or row["GMT Stok Adedi"] == 0:
        match_gmt = df_gmt_final[
            (df_gmt_final["GMT Ürün Kodu"].astype(str) == urunkodu) &
            (df_gmt_final["Firma Kodu"].astype(str)   == str(firma_kodu))
        ]
        if not match_gmt.empty:
            m_idx     = match_gmt.index[0]
            stok_val  = match_gmt.iloc[0]["GMT Stok Adedi"]
            df_calisma_alani.at[idx, "GMT Stok Adedi"] = "GMT'de Var" if stok_val > 0 else stok_val
            used_gmt_indices_step2.append(m_idx)

    # SİTA
    if pd.isna(row["SİTA Stok Adedi"]) or row["SİTA Stok Adedi"] == 0:
        match_sita = df_sita_final[
            (df_sita_final["SİTA Ürün Kodu"].astype(str) == urunkodu) &
            (df_sita_final["Firma Kodu"].astype(str)     == str(firma_kodu))
        ]
        if not match_sita.empty:
            m_idx     = match_sita.index[0]
            stok_val  = match_sita.iloc[0]["SİTA Stok Adedi"]
            df_calisma_alani.at[idx, "SİTA Stok Adedi"] = "SİTA'da Var" if stok_val > 0 else stok_val
            used_sita_indices_step2.append(m_idx)

df_gmt_final  = df_gmt_final.drop(used_gmt_indices_step2).reset_index(drop=True)
df_sita_final = df_sita_final.drop(used_sita_indices_step2).reset_index(drop=True)








# ------------------------------------------------------------
# 6) Tek Çıktı: "Nirvana.xlsx"
# ------------------------------------------------------------

def size_sort_key(s):
    """
    Beden değerlerinin sıralanması için özel anahtar.
    - Önce eğer tamamen sayısal ise numerik olarak sıralar (örneğin: 32, 34, 35).
    - Ardından '2XL', '3XL' gibi kalıpları, varsa karşılıklarını (XL, XXL, XXXL) kullanarak sıralar.
    - Eğer değer, ["XXS", "XS", "S", "M", "L", "XL", "XXL", "XXXL"] dizisindeyse bu sıraya göre sıralar.
    - Diğer durumlarda alfabetik sıralama kullanılır.
    """
    s_clean = s.strip().upper()
    s_numeric = s_clean.replace(" ", "")
    if s_numeric.isdigit():
        return (0, int(s_numeric))
    # '2XL', '3XL' gibi kalıplar için kontrol
    match = re.match(r'^(\d+)XL$', s_clean)
    if match:
        num = int(match.group(1))
        mapping = {1: "XL", 2: "XXL", 3: "XXXL"}
        common_order = ["XXS", "XS", "S", "M", "L", "XL", "XXL", "XXXL"]
        if num in mapping:
            return (1, common_order.index(mapping[num]))
        else:
            # Tanımlı aralık dışında kalanları büyük bir değerle sıralamanın sonuna alır
            return (1, 1000 + num)
    common_order = ["XXS", "XS", "S", "M", "L", "XL", "XXL", "XXXL"]
    if s_clean in common_order:
        return (1, common_order.index(s_clean))
    # Diğer tüm durumlarda alfabetik sıralama
    return (2, s_clean)

def temizle_ozellik_metni(value):
    """
    Verilen metni noktalı virgülle ayırır.
    - "Beden:" ile başlayan parçaları ayrıştırıp, Beden değerini metin içerisinden çıkarır.
      Bu değerler, sıralandıktan sonra "Asorti" adındaki kolonda saklanacaktır.
    - "Renk Seçiniz:" ve "Kategori Seçiniz:" ile başlayan parçalar temizlenen metinden atılır.
    - Geri kalan parçalar " // " ile birleştirilir.
    
    Fonksiyon, temizlenmiş metin ve asorti değerini içeren (cleaned_text, asorti_joined) şeklinde iki değer döndürür.
    """
    if not isinstance(value, str):
        return value  # Metin değilse dokunma

    # Noktalı virgülle parçala
    segments = [seg.strip() for seg in value.split(';')]
    
    # Temizleme sırasında atılacak anahtarlar (Beden dışındaki)
    remove_starts = ["Renk Seçiniz:", "Kategori Seçiniz:"]
    
    filtered_segments = []
    asorti_sizes = []
    
    for seg in segments:
        if seg.startswith("Beden:"):
            # "Beden:" ifadesini kaldırıp kalan kısmı al
            size_value = seg.replace("Beden:", "").strip()
            if size_value:
                asorti_sizes.append(size_value)
        elif any(seg.startswith(r) for r in remove_starts):
            # Belirlenen anahtarlarla başlayanları atla
            continue
        else:
            filtered_segments.append(seg)
    
    # Asorti beden değerlerini mantıklı sıraya göre sırala
    asorti_sizes_sorted = sorted(asorti_sizes, key=size_sort_key)
    asorti_joined = " // ".join(asorti_sizes_sorted)
    
    # Kalan metni " // " ile birleştir
    cleaned_text = " // ".join(filtered_segments)
    
    # Fonksiyon iki değeri tuple olarak döndürür:
    return cleaned_text, asorti_joined

# Örnek: DataFrame'in "Ozellik" kolonunu yeni işlev ile dönüştürüp hem temizlenmiş metni hem de asorti bedenleri ayırıyoruz.
# df_calisma_alani DataFrame'inizin daha önceden tanımlı olduğunu varsayıyoruz.
df_calisma_alani[["Ozellik", "Asorti"]] = df_calisma_alani["Ozellik"].apply(lambda x: pd.Series(temizle_ozellik_metni(x)))

# Ardından Excel'e kayıt işlemini yapıyoruz:
df_calisma_alani.to_excel("Nirvana.xlsx", index=False)
































# -----------------------------------------------------------
# 1) Nirvana.xlsx'i oku ve firma kodunu ayıkla
# -----------------------------------------------------------
def parse_firma_code(text):
    """
    Daha esnek regex:  \.(\d+)\.?$
    Örnek: 
      "Dabıl Kumaş Pantolon Siyah - 6003.1247." -> 1247
      "6003.1247" -> 1247
    """
    if not isinstance(text, str):
        return None
    text = text.strip()
    match = re.search(r'\.(\d+)\.?$', text)
    if match:
        return match.group(1)
    return None

nirvana_df = pd.read_excel("Nirvana.xlsx")
nirvana_df["FirmaKodu"] = nirvana_df["UrunAdi"].apply(parse_firma_code)

# -----------------------------------------------------------
# 2) API'ye bağlanarak vendor (sayfa sayfa 100 er kayıt)
#    ve "name" alanından kod ayıklama
# -----------------------------------------------------------
def login():
    conn = http.client.HTTPSConnection("task.haydigiy.com")

    login_payload = {
        "apiKey": "MypGcaEInEOTzuYQydgDHQ",
        "secretKey": "jRqliBLDPke76YhL_WL5qg",
        "emailOrPhone": "mustafa_kod@haydigiy.com",
        "password": "123456"
    }
    login_headers = {
        "Content-Type": "application/json",
        "Accept": "application/json"
    }

    conn.request("POST", "/api/customer/login", body=json.dumps(login_payload), headers=login_headers)
    res = conn.getresponse()
    data = res.read().decode("utf-8")
    if res.status != 200:
        print("Giriş başarısız:", data)
        return None

    login_data = json.loads(data)
    token = login_data.get("data", {}).get("token")
    if not token:
        print("Token alınamadı. Dönen veri:", login_data)
        return None

    return token

def get_all_vendor_data(token):
    """
    Tek seferde 100 kayıt dönebiliyor.
    Veriler bitene kadar pageIndex arttırarak tümünü toplayalım.
    """
    all_items = []
    page_index = 1
    while True:
        conn = http.client.HTTPSConnection("task.haydigiy.com")
        list_payload = {
            "manufacturerName": "",
            "published": None,
            "pageIndex": page_index,
            "pageSize": 100
        }
        list_headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
            "Accept": "application/json"
        }

        conn.request("POST", "/adminapi/vendor/list", body=json.dumps(list_payload), headers=list_headers)
        res = conn.getresponse()
        data = res.read().decode("utf-8")
        if res.status != 200:
            print(f"Sayfa {page_index} isteği başarısız:", data)
            break

        response_data = json.loads(data)
        items = response_data.get("data", [])
        if not items:
            # artık veri gelmiyorsa veya boşsa döngüden çık
            break

        all_items.extend(items)
        # 100'den az geldiyse son sayfayız demektir
        if len(items) < 100:
            break

        page_index += 1

    return all_items

def parse_vendor_code(text):
    """
    "name" alanı örnek: ".1806." -> 1806
    Regex:  \.(\d+)\.?$
    """
    if not isinstance(text, str):
        return None
    text = text.strip()
    match = re.search(r'\.(\d+)\.?$', text)
    if match:
        return match.group(1)
    return None

# Login ol ve tüm vendor verisini topla
token = login()
vendor_data_list = []
vendor_dict = {}

if token:
    vendor_data_list = get_all_vendor_data(token)
    # vendor_data_list içindeki her item'da "name" alanı olduğunu varsayıyoruz
    for item in vendor_data_list:
        name_text = item.get("name", "")
        code = parse_vendor_code(name_text)
        if code:
            vendor_dict.setdefault(code, []).append(name_text)

# -----------------------------------------------------------
# 3) Nirvana DataFrame'inde vendor kodlarını eşleştir
# -----------------------------------------------------------
def get_vendor_note_for_code(code):
    """
    Kod dictionary'de varsa, o koda bağlı metin(leri) "//" ile birleştirerek döndür.
    Yoksa "" döndür.
    """
    if not code:
        return ""
    return " // ".join(vendor_dict[code]) if code in vendor_dict else ""

nirvana_df["VendorNotu"] = nirvana_df["FirmaKodu"].apply(get_vendor_note_for_code)

# -----------------------------------------------------------
# 4) Ara dosya oluşturmadan doğrudan openpyxl ile Excel çıktısı
#    (DataFrame verisini tek tek hücrelere yazar + comment ekler)
# -----------------------------------------------------------

# Yeni bir workbook oluştur
wb = Workbook()
ws = wb.active

# 4A) Sütun başlıklarını yaz
columns = list(nirvana_df.columns)  # ['UrunAdi', 'StokAdedi', ..., 'FirmaKodu', 'VendorNotu', ...]
for col_index, col_name in enumerate(columns, start=1):
    ws.cell(row=1, column=col_index, value=col_name)

# 4B) DataFrame içeriğini hücrelere doldur
for row_index, row_data in nirvana_df.iterrows():
    excel_row = row_index + 2  # 1. satır başlık olduğu için
    for col_index, col_name in enumerate(columns, start=1):
        cell_value = row_data[col_name]
        cell = ws.cell(row=excel_row, column=col_index, value=cell_value)

# 4C) openpyxl ile comment ekle (örn. "UrunAdi" sütunu)
col_urunadi_idx = columns.index("UrunAdi") + 1  # 1-based
for row_index, row_data in nirvana_df.iterrows():
    vendor_note = row_data["VendorNotu"]
    if vendor_note.strip():
        # "UrunAdi" hücresine comment ekle
        excel_row = row_index + 2
        urun_adi_cell = ws.cell(row=excel_row, column=col_urunadi_idx)
        urun_adi_cell.comment = Comment(text=vendor_note, author="VendorInfo")

# 4D) Son olarak kaydet
wb.save("Nirvana.xlsx")
































# ------------------------------------------------------------
# EK ADIMLAR
# ------------------------------------------------------------
# 1) "ListeFiyatı" sütunu oluşturma (openpyxl ile)
#    Bu aşamada 'Kar Yüzdesi', 'Alış Fiyatı', 'Kategori' kolonlarının
#    zaten var olduğunu varsayıyoruz.



dosya_adi = "Nirvana.xlsx"
workbook = openpyxl.load_workbook(dosya_adi)
kopya_sayfa_adi = "Sheet1"  # Kendi sayfa adınıza göre değiştirebilirsiniz

if kopya_sayfa_adi in workbook.sheetnames:
    sheet = workbook[kopya_sayfa_adi]
    
    # Başlıkları okuyup (ilk satır) sütun isimlerini ve indekslerini sözlüğe atıyoruz
    basliklar = {cell.value: cell.column for cell in sheet[1] if cell.value}

    # "VaryasyonGittiGidiyorKodu" sütununu bul
    vgk_kolon = basliklar.get("VaryasyonGittiGidiyorKodu")

    if vgk_kolon:
        # 1) En sağdaki ilk boş sütunu belirleyelim
        yeni_sutun_index = sheet.max_column + 1  # max_column'ın bir sonrasına eklenecek

        # 2) Yeni sütuna, "VaryasyonGittiGidiyorKodu" değerlerini kopyalayalım (opsiyonel)
        for row in range(1, sheet.max_row + 1):
            eski_hucre = sheet.cell(row=row, column=vgk_kolon)
            yeni_hucre = sheet.cell(row=row, column=yeni_sutun_index)
            yeni_hucre.value = eski_hucre.value
            # Stil kopyalama (opsiyonel)
            if eski_hucre.has_style:
                yeni_hucre._style = copy(eski_hucre._style)

        # 3) Yeni sütunun başlığını "Kar Yüzdesi" yapalım (1. satır)
        sheet.cell(row=1, column=yeni_sutun_index).value = "Kar Yüzdesi"

        # 4) Satış Fiyatı ve Alış Fiyatı sütunlarının indekslerini alalım
        satis_fiyati_kolon = basliklar.get("SatisFiyati")
        alis_fiyati_kolon = basliklar.get("AlisFiyati")

        # 5) Her satır için kâr yüzdesi hesapla
        #    (Kar yüzdesi = (Satış - Alış) / Satış) * 100
        if satis_fiyati_kolon and alis_fiyati_kolon:
            for row in range(2, sheet.max_row + 1):
                satis_fiyati = sheet.cell(row=row, column=satis_fiyati_kolon).value
                alis_fiyati = sheet.cell(row=row, column=alis_fiyati_kolon).value

                # "Kar Yüzdesi" hücresi (yeni sütun)
                kar_yuzdesi_hucre = sheet.cell(row=row, column=yeni_sutun_index)

                if (satis_fiyati is not None) and (alis_fiyati is not None):
                    try:
                        kar_yuzdesi = (satis_fiyati - alis_fiyati) / satis_fiyati * 100
                        kar_yuzdesi_hucre.value = kar_yuzdesi
                        kar_yuzdesi_hucre.number_format = "0.00"
                    except ZeroDivisionError:
                        kar_yuzdesi_hucre.value = None

        # Değişiklikleri kaydedip workbook'u kapatalım
        workbook.save(dosya_adi)
        gc.collect()


# ============================================================
# 2) "ListeFiyatı" kolonunu oluşturma
# ============================================================
workbook = openpyxl.load_workbook(dosya_adi)

# Açılacak sayfa ismi ("Sheet1" veya oluşturulan tablo ismi olabilir).
sheet_adi = workbook.active.title  # Burada direkt aktif sayfanın adını aldık
# İsterseniz sheet_adi = "Sheet1" şeklinde de doğrudan belirtebilirsiniz.

if sheet_adi in workbook.sheetnames:
    sheet = workbook[sheet_adi]

    # Başlık hücrelerini (ilk satır) bir sözlükte tutuyoruz
    basliklar = {cell.value: cell.column for cell in sheet[1] if cell.value}

    kar_yuzdesi_kolon = basliklar.get("Kar Yüzdesi")
    alis_fiyati_kolon = basliklar.get("AlisFiyati")
    kategori_kolon = basliklar.get("Kategori")

    if kar_yuzdesi_kolon and alis_fiyati_kolon and kategori_kolon:
        # Yeni sütunu, "Kar Yüzdesi" sütununun hemen sağ tarafına ekleyeceğiz
        yeni_kolon_index = kar_yuzdesi_kolon + 1

        # 1) Kar Yüzdesi sütunundaki hücreleri biçimiyle birlikte yeni sütuna kopyalayalım
        for row in range(1, sheet.max_row + 1):
            eski_hucre = sheet.cell(row=row, column=kar_yuzdesi_kolon)
            yeni_hucre = sheet.cell(row=row, column=yeni_kolon_index)
            yeni_hucre.value = eski_hucre.value
            if eski_hucre.has_style:
                yeni_hucre._style = copy(eski_hucre._style)

        # 2) Yeni sütunun başlığını "ListeFiyatı" yap
        sheet.cell(row=1, column=yeni_kolon_index).value = "ListeFiyatı"

        # 3) Her satır için alış fiyatına göre "Liste Fiyatı" hesapla
        for row in range(2, sheet.max_row + 1):
            alis_fiyati = sheet.cell(row=row, column=alis_fiyati_kolon).value
            kategori = sheet.cell(row=row, column=kategori_kolon).value
            yeni_hucre = sheet.cell(row=row, column=yeni_kolon_index)

            if alis_fiyati is not None:
                # Alış fiyatına göre temel tutar
                if 0 <= alis_fiyati <= 24.99:
                    result = alis_fiyati + 10
                elif 25 <= alis_fiyati <= 39.99:
                    result = alis_fiyati + 13
                elif 40 <= alis_fiyati <= 59.99:
                    result = alis_fiyati + 17
                elif 60 <= alis_fiyati <= 200.99:
                    result = alis_fiyati * 1.40
                elif alis_fiyati >= 201:
                    result = alis_fiyati * 1.35
                else:
                    result = alis_fiyati

                # Kategori takı/parfüm/aksesuar vs. ise ekstra çarpan
                if (
                    isinstance(kategori, str) and 
                    any(cat in kategori for cat in ["Parfüm", "Gözlük", "Saat", "Kolye", "Küpe", "Bileklik", "Bilezik"])
                ):
                    result *= 1.20
                else:
                    result *= 1.10

                # Son olarak tam sayıya yuvarlayıp 0.99 ekleyelim
                result = int(round(result)) + 0.99

                # Para formatı (₺) ekle
                yeni_hucre.value = result
                yeni_hucre.number_format = '#,##0.00₺'

workbook.save(dosya_adi)
gc.collect()


# ------------------------------------------------------------
# 2) "AramaTerimleri" kolonu için tarih ayıklama
# ------------------------------------------------------------

date_pattern = r'(\d{1,2}\.\d{1,2}\.\d{4})'

def extract_date(x):
    match = re.search(date_pattern, str(x))
    if match:
        # Elde edilen tarih metnini datetime objesine çeviriyoruz
        return pd.to_datetime(match.group(1), format='%d.%m.%Y')
    return pd.NaT  # Tarih bulunamazsa Not-a-Time döndür

df_calisma_alani['AramaTerimleri'] = df_calisma_alani['AramaTerimleri'].apply(extract_date)




# İlk olarak mevcut "Kategori" kolonunu "Özel Kategoriler" adlı yeni bir kolona kopyala
df_calisma_alani['Özel Kategoriler'] = df_calisma_alani['Kategori']

# ------------------------------------------------------------
# 3) "Kategori" kolonunda düzenleme
# ------------------------------------------------------------
df_calisma_alani['Kategori'] = df_calisma_alani['Kategori'].fillna("")

import re

def extract_category(text):
    if not isinstance(text, str):
        return None
    
    # Başta "Butik" kontrolü
    if "Butik" in text:
        return "Butik"
    
    # 1) Önce TESETTÜR kontrolü
    if "TESETTÜR" in text:
        return "TESETTÜR"
    
    # 2) ";" üzerinden parçalara ayır
    parts = text.split(';')
    
    # 3) Her parçadaki kategori metnini ">" ya da ">>" den sonra al
    categories = []
    for part in parts:
        sub_parts = re.split(r'>>|>', part)
        if len(sub_parts) > 1:
            cat = sub_parts[-1].strip()
            categories.append(cat)
    
    # 4) Tek kategori geldiyse o kategoriyi döndür
    if len(categories) == 1:
        return categories[0]
    
    # 5) Birden çok kategori varsa
    if len(categories) > 1:
        normal_cats = [cat for cat in categories if "Büyük Beden" not in cat]
        bb_cats = [cat for cat in categories if "Büyük Beden" in cat]
        # Eğer hem normal kategori hem de "Büyük Beden" içeren kategori varsa ikisini birleştir
        if normal_cats and bb_cats:
            return f"{normal_cats[0]} // {bb_cats[0]}"
        # Sadece normal kategori varsa
        if normal_cats:
            return normal_cats[0]
        # Sadece "Büyük Beden" içeren kategori varsa
        if bb_cats:
            return bb_cats[0]
    
    # Hiç kategori bulunamadıysa None döndür
    return None

df_calisma_alani['Kategori'] = df_calisma_alani['Kategori'].apply(extract_category)


# ------------------------------------------------------------
# 4) Yeni kolonlar: "Stok Adedi Her Şey Dahil" ve "Stok Adedi Site ve Vega"
# ------------------------------------------------------------
# Burada kullanılan kolon adlarına dikkat: 
# "GMT Stok Adedi", "SİTA Stok Adedi", "StokAdedi" ve "VaryasyonMorhipoKodu" 
# yoksa oluşturulabilir.

if "GMT Stok Adedi" not in df_calisma_alani.columns:
    df_calisma_alani["GMT Stok Adedi"] = 0
if "SİTA Stok Adedi" not in df_calisma_alani.columns:
    df_calisma_alani["SİTA Stok Adedi"] = 0
if "StokAdedi" not in df_calisma_alani.columns:
    df_calisma_alani["StokAdedi"] = 0
if "VaryasyonMorhipoKodu" not in df_calisma_alani.columns:
    df_calisma_alani["VaryasyonMorhipoKodu"] = 0

gmt_numeric = pd.to_numeric(df_calisma_alani["GMT Stok Adedi"], errors="coerce").fillna(0)
sita_numeric = pd.to_numeric(df_calisma_alani["SİTA Stok Adedi"], errors="coerce").fillna(0)
stok_adedi_numeric = pd.to_numeric(df_calisma_alani["StokAdedi"], errors="coerce").fillna(0)
n11_zimmet_numeric = pd.to_numeric(df_calisma_alani["VaryasyonMorhipoKodu"], errors="coerce").fillna(0)

df_calisma_alani["Stok Adedi Her Şey Dahil"] = stok_adedi_numeric + n11_zimmet_numeric + gmt_numeric + sita_numeric
df_calisma_alani["Stok Adedi Site ve Vega"] = stok_adedi_numeric + n11_zimmet_numeric

df_calisma_alani['StokAdedi'].fillna(0, inplace=True)
df_calisma_alani['VaryasyonMorhipoKodu'].fillna(0, inplace=True)
df_calisma_alani['GMT Stok Adedi'].fillna(0, inplace=True)
df_calisma_alani['SİTA Stok Adedi'].fillna(0, inplace=True)

# ------------------------------------------------------------
# 5) Yeni kolonlar: "Kaç Güne Biter Her Şey Dahil" ve "Kaç Güne Biter Site ve Vega"
#    Hesaplama: Stok / MorhipoKodu
# ------------------------------------------------------------
# Kolon isimleri tutarlılık için kontrol ediliyor
if "MorhipoKodu" not in df_calisma_alani.columns:
    df_calisma_alani["MorhipoKodu"] = 0

df_calisma_alani['MorhipoKodu'].fillna(0, inplace=True)
df_calisma_alani["Kaç Güne Biter Her Şey Dahil"] = "Satış Adedi Yok"
df_calisma_alani["Kaç Güne Biter Site ve Vega"] = "Satış Adedi Yok"

non_zero_mask = df_calisma_alani["MorhipoKodu"] != 0
df_calisma_alani.loc[non_zero_mask, "Kaç Güne Biter Her Şey Dahil"] = round(
    df_calisma_alani["Stok Adedi Her Şey Dahil"] / df_calisma_alani["MorhipoKodu"]
)

df_calisma_alani.loc[non_zero_mask, "Kaç Güne Biter Site ve Vega"] = round(
    df_calisma_alani["Stok Adedi Site ve Vega"] / df_calisma_alani["MorhipoKodu"]
)

# ------------------------------------------------------------
# 6) "Görüntülenmenin Satışa Dönüş Oranı" kolonu
#    "HepsiBuradaKodu" -> "Ortalama Görüntülenme Adedi" rename
# ------------------------------------------------------------
# Kolon yoksa ekliyoruz
if "HepsiBuradaKodu" not in df_calisma_alani.columns:
    df_calisma_alani["HepsiBuradaKodu"] = 0

df_calisma_alani = df_calisma_alani.rename(columns={"HepsiBuradaKodu": "Ortalama Görüntülenme Adedi"})
df_calisma_alani['Ortalama Görüntülenme Adedi'].fillna(0, inplace=True)

df_calisma_alani["Görüntülenmenin Satışa Dönüş Oranı"] = "0"

non_zero_mask = df_calisma_alani["Ortalama Görüntülenme Adedi"] != 0
df_calisma_alani.loc[non_zero_mask, "Görüntülenmenin Satışa Dönüş Oranı"] = round(
    (df_calisma_alani["MorhipoKodu"] / df_calisma_alani["Ortalama Görüntülenme Adedi"]) * 100, 2
)

# Tekrar kaydediyoruz
df_calisma_alani.to_excel("Nirvana.xlsx", index=False)























def duzenleme_islemleri(dosya_adi="Nirvana.xlsx", sayfa_adi="Sheet1"):
    # 1) Excel'i aç
    workbook = openpyxl.load_workbook(dosya_adi)
    if sayfa_adi not in workbook.sheetnames:
        print(f"'{sayfa_adi}' isminde bir sayfa bulunamadı.")
        return
    sheet = workbook[sayfa_adi]

    # İlk satırdan başlık -> sütun indekslerini elde et
    basliklar = {}
    for cell in sheet[1]:
        if cell.value:  # None değilse
            basliklar[cell.value] = cell.column

    # ------------------------------------------------
    # 1) "VaryasyonGittiGidiyorKodu" sütununu en sağdaki ilk boş sütuna kopyala,
    #    yeni sütuna "Kar Yüzdesi" adını ver + formül uygula
    # ------------------------------------------------
    vgk_baslik = "VaryasyonGittiGidiyorKodu"
    if vgk_baslik in basliklar:
        # "VaryasyonGittiGidiyorKodu" kolonunu bul
        stok_adedi_kolon = basliklar[vgk_baslik]
        # Yeni sütun yeri: En sağdaki ilk boş sütun
        yeni_sutun_index = sheet.max_column + 1

        # Eski hücre -> yeni hücre kopyalama (değer + stil)
        for row in range(1, sheet.max_row + 1):
            eski_hucre = sheet.cell(row=row, column=stok_adedi_kolon)
            yeni_hucre = sheet.cell(row=row, column=yeni_sutun_index)
            yeni_hucre.value = eski_hucre.value
            if eski_hucre.has_style:
                yeni_hucre._style = copy(eski_hucre._style)

        # Yeni sütuna başlık ismi
        sheet.cell(row=1, column=yeni_sutun_index).value = "Kar Yüzdesi"

        # Başlıkları tekrar güncelle (yeni sütun eklendi)
        basliklar = {}
        for cell in sheet[1]:
            if cell.value:
                basliklar[cell.value] = cell.column

        # "Satış Fiyatı" ve "Alış Fiyatı" sütun indekslerini al
        satis_fiyati_kolon = basliklar.get("SatisFiyati")
        alis_fiyati_kolon = basliklar.get("AlisFiyati")
        kar_yuzdesi_kolon = basliklar.get("Kar Yüzdesi")

        if satis_fiyati_kolon and alis_fiyati_kolon and kar_yuzdesi_kolon:
            for row in range(2, sheet.max_row + 1):
                sf = sheet.cell(row=row, column=satis_fiyati_kolon).value
                af = sheet.cell(row=row, column=alis_fiyati_kolon).value
                kar_hucre = sheet.cell(row=row, column=kar_yuzdesi_kolon)

                if sf and af:
                    try:
                        kar_orani = (sf - af) / sf
                        kar_hucre.value = kar_orani
                        kar_hucre.number_format = "0.00%"
                    except ZeroDivisionError:
                        kar_hucre.value = None

    # ------------------------------------------------
    # "Üründe Hareket Var mı?" adında yeni bir kolon oluşturma
    # ------------------------------------------------

    # ------------------------------------------------
    # 2) "Üründe Hareket Var mı?" adında yeni bir kolon oluştur
    # ------------------------------------------------
    yeni_kolon_sira = sheet.max_column + 1
    sheet.cell(row=1, column=yeni_kolon_sira).value = "Üründe Hareket Var mı?"

    # Başlıkları güncelleyelim
    basliklar = {}
    for cell in sheet[1]:
        if cell.value:  # None değilse
            basliklar[cell.value] = cell.column

    # İlgili kolon indeksleri
    resim_kolon = basliklar.get("Resim")
    stok_hersey_kolon = basliklar.get("Stok Adedi Her Şey Dahil")
    gunluk_ort_satis_kolon = basliklar.get("MorhipoKodu")
    dunun_satis_kolon = basliklar.get("Adet")

    for row in range(2, sheet.max_row + 1):
        sonuc_hucre = sheet.cell(row=row, column=yeni_kolon_sira)
        
        # 1) RESİM KONTROLÜ
        resim_deger = sheet.cell(row=row, column=resim_kolon).value if resim_kolon else None
        if not resim_deger or resim_deger == 0:
            sonuc_hucre.value = "Resim Yok"
            continue

        # 2) İLGİLİ KOLONLARI AL
        stok_deger = sheet.cell(row=row, column=stok_hersey_kolon).value if stok_hersey_kolon else 0
        gunluk_ort_satis = sheet.cell(row=row, column=gunluk_ort_satis_kolon).value if gunluk_ort_satis_kolon else 0
        dunun_satis = sheet.cell(row=row, column=dunun_satis_kolon).value if dunun_satis_kolon else 0
        
        # None gelirse 0 kabul edelim
        stok_deger = stok_deger if stok_deger else 0
        gunluk_ort_satis = gunluk_ort_satis if gunluk_ort_satis else 0
        dunun_satis = dunun_satis if dunun_satis else 0
        
        # 3) EVET KONTROLÜ
        # "Evet" => Günlük Ortalama Satış Adedi > 0 veya Dünün Satış Adedi > 0
        if (gunluk_ort_satis > 0) or (dunun_satis > 0) or (stok_deger > 0):
            sonuc_hucre.value = "Evet"
        
        # 4) HAYIR KONTROLÜ
        # "Hayır" => Stok <= 0 VE Günlük Ortalama Satış Adedi <= 0 VE Dünün Satış Adedi <= 0
        elif (stok_deger <= 0) and (gunluk_ort_satis <= 0) and (dunun_satis <= 0):
            sonuc_hucre.value = "Hayır"
        
        # 5) EĞER HİÇBİRİ TUTMADIYSA
        else:
            sonuc_hucre.value = "Hayır"
    # ------------------------------------------------
    # 3) TrendyolKodu ve VaryasyonTrendyolKodu kolonlarındaki
    #    veriler için ilk boşluktan sonraki kısmı temizle
    # ------------------------------------------------
    trendyol_kodu_baslik = "TrendyolKodu"
    varyant_trendyol_kodu_baslik = "VaryasyonTrendyolKodu"

    def remove_after_first_space(text):
        """
        "ABC DEF GHI" -> "ABC"
        İçerik boş veya None ise dokunma.
        """
        if not text or not isinstance(text, str):
            return text
        # En fazla 1 kez bölelim
        parts = text.split(" ", 1)
        return parts[0] if len(parts) > 1 else parts[0]

    for col_name in [trendyol_kodu_baslik, varyant_trendyol_kodu_baslik]:
        if col_name in basliklar:
            c_index = basliklar[col_name]
            for row in range(2, sheet.max_row + 1):
                hucre = sheet.cell(row=row, column=c_index)
                hucre.value = remove_after_first_space(hucre.value)

    # ------------------------------------------------
    # 4) İstenilen kolonları silelim + Kolon isimlerini değiştirelim
    # ------------------------------------------------

    # 4.2) "ToplamFiyat" kolonunu silelim (eğer varsa)
    if "ToplamFiyat" in basliklar:
        col_idx = basliklar["ToplamFiyat"]
        sheet.delete_cols(col_idx, 1)
        basliklar = {}
        for cell in sheet[1]:
            if cell.value:
                basliklar[cell.value] = cell.column

    # 4.3) Kolonları tekrar güncelleyip yeni isimler atayalım
    rename_map = {
        "UrunAdi": "Ürün Adı",
        "AlisFiyati": "Alış Fiyatı",
        "SatisFiyati": "Satış Fiyatı",
        "AramaTerimleri": "Resim Yüklenme Tarihi",
        "MorhipoKodu": "Günlük Ortalama Satış Adedi",
        "VaryasyonMorhipoKodu": "Depodaki Adetler",
        "N11Kodu": "Mevsim",
        "VaryasyonGittiGidiyorKodu": "Net Satış Tarihi ve Adedi",
        "TrendyolKodu": "Son Transfer Tarihi",
        "VaryasyonTrendyolKodu": "Son İndirim Tarihi",
        "StokAdedi": "Instagram Stok Adedi",
        "Adet": "Dünün Satış Adedi",
        "ListeFiyatı": "Liste Fiyatı",
        "Ozellik": "Etiketler",
        "Varyasyon": "Bedenler",
        "VaryasyonAmazonKodu": "Sigara Ürün Ekle / Çıkar"
    }

    for eski, yeni in rename_map.items():
        if eski in basliklar:
            col_idx = basliklar[eski]
            sheet.cell(row=1, column=col_idx).value = yeni

    # Sütunları yeniden oku
    basliklar = {}
    for cell in sheet[1]:
        if cell.value:
            basliklar[cell.value] = cell.column

    # ------------------------------------------------
    # 5) Son olarak, sütunların sırasını yeniden düzenle
    # ------------------------------------------------
    final_order = [
        "StokKodu",
        "Ürün Adı",
        "Üründe Hareket Var mı?",
        "Instagram Stok Adedi",
        "Bedenler",
        "Stok Adedi Her Şey Dahil",
        "Stok Adedi Site ve Vega",
        "Günlük Ortalama Satış Adedi",  
        "Dünün Satış Adedi",  
        "Kaç Güne Biter Her Şey Dahil",
        "Net Satış Tarihi ve Adedi",         
        "Ortalama Görüntülenme Adedi", 
        "Görüntülenmenin Satışa Dönüş Oranı",
        "Kaç Güne Biter Site ve Vega",
        "Resim",
        "Alış Fiyatı",
        "Son Satın Alma Fiyatı",
        "Satış Fiyatı",
        "Liste Fiyatı",
        "Kar Yüzdesi",
        "Resim Yüklenme Tarihi",
        "Kategori",
        "GMT Stok Adedi",
        "SİTA Stok Adedi",
        "Mevsim",
        "Son Transfer Tarihi",
        "Son İndirim Tarihi",
        "Marka",
        "Etiketler",
        "Asorti",
        "Özel Kategoriler",
        "Sigara Ürün Ekle / Çıkar"
    ]

    # Mevcut tüm veriyi memory'e alıyoruz (list of dict)
    all_data = []
    headers_in_sheet = [cell.value for cell in sheet[1] if cell.value]

    for r in range(2, sheet.max_row + 1):
        row_dict = {}
        for header in headers_in_sheet:
            col_idx = basliklar.get(header)
            if col_idx:
                row_dict[header] = sheet.cell(row=r, column=col_idx).value
            else:
                row_dict[header] = None
        all_data.append(row_dict)

    # Sayfayı temizle (başlık satırı da dahil)
    sheet.delete_rows(1, sheet.max_row)

    # Yeni başlık satırını final_order'a göre yaz
    for col_idx, header_name in enumerate(final_order, start=1):
        sheet.cell(row=1, column=col_idx).value = header_name

    # all_data'daki her satır için final_order sırasına göre yaz
    for row_idx, row_data in enumerate(all_data, start=2):
        for col_idx, header_name in enumerate(final_order, start=1):
            value = row_data.get(header_name, None)
            sheet.cell(row=row_idx, column=col_idx).value = value

    # ------------------------------------------------------------
    # 6) En son: "Resim" kolonunda .jpeg kısaltma + "Ürün Adı" hücresine link ekle
    # ------------------------------------------------------------
    # Çünkü üstteki reorder işlemi hyperlink bilgisini siliyor.
    # Bu yüzden hyperlink atamayı en sona koyuyoruz.

    # Başlıkları tekrar toplayalım, zira rename oldu ("UrunAdi" -> "Ürün Adı" vb.)
    basliklar_son = {}
    for cell in sheet[1]:
        if cell.value:
            basliklar_son[cell.value] = cell.column

    # "Resim" -> "Ürün Adı"
    resim_header_son = "Resim"
    urun_adi_header_son = "Ürün Adı"

    if resim_header_son in basliklar_son and urun_adi_header_son in basliklar_son:
        resim_col_idx = basliklar_son[resim_header_son]
        urun_adi_col_idx = basliklar_son[urun_adi_header_son]

        for row in range(2, sheet.max_row + 1):
            resim_deger = sheet.cell(row=row, column=resim_col_idx).value
            if isinstance(resim_deger, str) and resim_deger.strip():
                # ";" varsa parçala, ilk linki al
                link_listesi = [lnk.strip() for lnk in resim_deger.split(';') if lnk.strip()]
                if link_listesi:
                    first_link = link_listesi[0]
                    # .jpeg sonrası silinsin
                    idx = first_link.lower().find(".jpeg")
                    truncated_value = first_link
                    if idx != -1:
                        idx += len(".jpeg")
                        truncated_value = first_link[:idx]

                    # Resim hücresine kısaltılmış değeri yaz
                    sheet.cell(row=row, column=resim_col_idx).value = truncated_value

                    # Kısaltılmış link http ile başlıyorsa "Ürün Adı" hücresine hyperlink
                    if truncated_value.lower().startswith("http"):
                        urun_adi_hucre = sheet.cell(row=row, column=urun_adi_col_idx)
                        # Aynı hücredeki metni koruyor, sadece link ekliyoruz.
                        urun_adi_hucre.hyperlink = truncated_value
                        # Not: Stil vermek istemiyorsak eklemeden geçiyoruz.

    # ------------------------------------------------
    # 7) Kaydet ve hafızayı temizle
    # ------------------------------------------------
    workbook.save(dosya_adi)
    gc.collect()



# Fonksiyonu çağır
if __name__ == "__main__":
    duzenleme_islemleri()

























def tablo_duzenleme(dosya_adi="Nirvana.xlsx", sayfa_adi="Sheet1"):
    # 1) Excel'i aç
    wb = openpyxl.load_workbook(dosya_adi)
    
    if sayfa_adi not in wb.sheetnames:
        print(f"'{sayfa_adi}' isminde bir sayfa bulunamadı.")
        return
    
    sheet = wb[sayfa_adi]

    # ------------------------------------------------
    # 1) "UrunKodu" kolonunu hariç tüm verileri yatayda ortala
    # ------------------------------------------------
    basliklar = {}
    for cell in sheet[1]:
        if cell.value:
            basliklar[cell.value] = cell.column

    urun_kodu_kolon_index = basliklar.get("UrunKodu")

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for cell_obj in row:
            if cell_obj.column != urun_kodu_kolon_index:
                cell_obj.alignment = Alignment(horizontal='center')

    # ------------------------------------------------
    # 2) Başlık satırının yüksekliğini 40 px yap
    # ------------------------------------------------
    sheet.row_dimensions[1].height = 40

    # ------------------------------------------------
    # 3) Başlık satırındaki verileri kalın, metni kaydırmalı,
    #    hem yatay hem dikey ortalı yapalım
    # ------------------------------------------------
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # ------------------------------------------------
    # 4) Freeze Panes: İlk satırı dondur
    # ------------------------------------------------
    sheet.freeze_panes = "A2"

    # ------------------------------------------------
    # 5) Sütun genişliklerini ayarla:
    #    - En az 120 px (yaklaşık 17.14 karakter genişliği)
    #    - Kolondaki en geniş veriye göre büyüt (karakter sayısı)
    # ------------------------------------------------
    min_width_chars = 120 / 7  # ~7 piksel/karakter
    for col in sheet.iter_cols(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row):
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        
        # Başlık satırını (col[0]) atlayarak veri uzunluğunu ölç
        for cell_obj in col[1:]:
            value = cell_obj.value
            if value is not None:
                length = len(str(value))
                if length > max_length:
                    max_length = length

        # Sütun genişliğini belirle
        optimal_width = max(min_width_chars, max_length + 2)
        sheet.column_dimensions[col_letter].width = optimal_width

    # ------------------------------------------------
    # 6) "Marka" kolonunda "Sigara Ürün" içeren satırları Açık Yeşil yap
    # ------------------------------------------------
    marka_col_index = basliklar.get("Marka")
    if marka_col_index:
        max_col = sheet.max_column
        yesil_fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")

        for row_idx in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_idx, column=marka_col_index).value
            if isinstance(cell_value, str) and "Sigara Ürün" in cell_value:
                for col_idx in range(1, max_col + 1):
                    sheet.cell(row=row_idx, column=col_idx).fill = yesil_fill

    # ------------------------------------------------
    # 7) Tabloyu "Beyaz Tablo Stili Açık 1" ile stillendirelim
    #    (bantlı satırlar için showRowStripes=True)
    # ------------------------------------------------
    max_row = sheet.max_row
    max_col = sheet.max_column
    start_col = get_column_letter(1)
    end_col = get_column_letter(max_col)
    tablo_araligi = f"{start_col}1:{end_col}{max_row}"

    my_table = Table(displayName="Tablom", ref=tablo_araligi)

    style = TableStyleInfo(
        name="TableStyleLight1", 
        showRowStripes=True,   # Şeritli (bantlı) satırlar
        showColumnStripes=False
    )
    my_table.tableStyleInfo = style
    sheet.add_table(my_table)

    # ------------------------------------------------
    # 8) Tüm kenarlıklar (thin) ekle
    #    Tablo aralığındaki her hücreyi ince siyah çerçeve yapıyoruz
    # ------------------------------------------------
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    for row_cells in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell_obj in row_cells:
            cell_obj.border = thin_border

    # ------------------------------------------------
    # 9) Değişiklikleri kaydet
    # ------------------------------------------------
    wb.save(dosya_adi)


if __name__ == "__main__":
    tablo_duzenleme()





def hide_columns_by_header(sheet, headers_to_hide):
    """
    Verilen sheet'te, birinci satır başlığı headers_to_hide listesinde
    varsa o kolonu gizler.
    """
    max_col = sheet.max_column
    for col_index in range(1, max_col + 1):
        header_value = sheet.cell(row=1, column=col_index).value
        if header_value in headers_to_hide:
            col_letter = get_column_letter(col_index)
            sheet.column_dimensions[col_letter].hidden = True

def copy_sheet_style_structure(source_sheet, target_sheet):
    """
    Kaynaktaki sütun genişlikleri, satır yükseklikleri ve
    birleştirilmiş hücreleri (merges) hedef sayfaya kopyalar.
    """
    # 1) Sütun genişliklerini kopyala
    for col_letter, col_dim in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[col_letter].width = col_dim.width

    # 2) Satır yüksekliklerini kopyala
    for row_idx, row_dim in source_sheet.row_dimensions.items():
        target_sheet.row_dimensions[row_idx].height = row_dim.height

    # 3) Merge (birleştirme) aralıklarını kopyala
    for merged_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_range))

def copy_row_with_style(source_sheet, target_sheet, src_row_idx, tgt_row_idx, max_col):
    """
    source_sheet'in src_row_idx satırını,
    target_sheet'in tgt_row_idx satırına, hücre stilleriyle birlikte kopyalar.
    """
    for c in range(1, max_col + 1):
        source_cell = source_sheet.cell(row=src_row_idx, column=c)
        target_cell = target_sheet.cell(row=tgt_row_idx, column=c)

        # Hücre değerini kopyala
        target_cell.value = source_cell.value

        # Stil kopyası (font, fill, border, alignment, number_format vb.)
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

def freeze_header(sheet):
    """
    Sayfanın ilk satırını (yani başlık satırını) dondurmak için 
    A2 hücresini 'freeze_panes' olarak ayarlar.
    """
    sheet.freeze_panes = "A2"

def apply_table_format(sheet, table_name="MyTable", style_name="TableStyleMedium9"):
    """
    Sayfadaki (A1'den başlayarak) mevcut veri aralığına 
    bir "Excel Tablosu" (Format as Table) ekler.
    Stil olarak 'TableStyleMedium9' kullanır 
    (veya isterseniz başka bir stil adını verebilirsiniz).
    
    - table_name: Tablonun workbook içindeki benzersiz adı.
    - style_name: "TableStyleMediumX", "TableStyleLightX" vs.
    """
    max_row = sheet.max_row
    max_col = sheet.max_column

    if max_row < 1 or max_col < 1:
        # Boş sayfa ise tablo oluşturulamaz
        return

    first_cell = "A1"
    last_cell = f"{get_column_letter(max_col)}{max_row}"
    table_ref = f"{first_cell}:{last_cell}"

    table = Table(displayName=table_name, ref=table_ref)
    style = TableStyleInfo(
        name=style_name,
        showRowStripes=True,   # Satırlar çizgili olsun
        showColumnStripes=False
    )
    table.tableStyleInfo = style

    sheet.add_table(table)

def filter_indirim_sheet_with_styles(
    source_sheet,
    target_sheet,
    instagram_stok_header="Instagram Stok Adedi",
    resim_header="Resim",
    urun_adi_header="Ürün Adı"
):
    """
    Özel olarak 'İndirim Raporu' sayfasının filtrelenmiş kopyasını oluşturur:
    - 'Instagram Stok Adedi' > 0
    - 'Resim' kolonundaki hücre boş olmayan satırları geçir
    - Her satırı kopyalarken stil bilgilerini de koru.
    - Kopyalama sonunda, 'Ürün Adı' kolonuna 'Resim' kolonundaki değeri hyperlink olarak ekle.
    """
    max_row = source_sheet.max_row
    max_col = source_sheet.max_column

    # Header (ilk satır)
    header_values = [source_sheet.cell(row=1, column=c).value for c in range(1, max_col+1)]

    # İlgili kolon indekslerini bulalım (1-bazlı)
    try:
        instagram_col_idx = header_values.index(instagram_stok_header) + 1
    except ValueError:
        instagram_col_idx = None

    try:
        resim_col_idx = header_values.index(resim_header) + 1
    except ValueError:
        resim_col_idx = None

    try:
        urun_adi_col_idx = header_values.index(urun_adi_header) + 1
    except ValueError:
        urun_adi_col_idx = None

    # Sayfanın sütun/row/merge yapılarını kopyala
    copy_sheet_style_structure(source_sheet, target_sheet)

    target_row_idx = 1
    for r in range(1, max_row + 1):
        if r == 1:
            # Header'ı (birinci satırı) doğrudan kopyala
            copy_row_with_style(source_sheet, target_sheet, r, target_row_idx, max_col)
            target_row_idx += 1
        else:
            # Filtre kontrolü
            pass_filter = True

            # 1) Instagram Stok Adedi > 0 mı?
            if instagram_col_idx is not None:
                val_instagram = source_sheet.cell(row=r, column=instagram_col_idx).value
                try:
                    if float(val_instagram) <= 0:
                        pass_filter = False
                except:
                    pass_filter = False  # sayısal değilse de at

            # 2) Resim kolonunda değer boş mu?
            if resim_col_idx is not None and pass_filter:
                val_resim = source_sheet.cell(row=r, column=resim_col_idx).value
                # Boş (None veya "") ise atla
                if val_resim is None or str(val_resim).strip() == "":
                    pass_filter = False

            if pass_filter:
                # Satırı kopyala
                copy_row_with_style(source_sheet, target_sheet, r, target_row_idx, max_col)

                # Ek olarak, "Ürün Adı" hücresine hyperlink verelim
                if (resim_col_idx is not None) and (urun_adi_col_idx is not None):
                    link_value = source_sheet.cell(row=r, column=resim_col_idx).value
                    product_name_cell = target_sheet.cell(row=target_row_idx, column=urun_adi_col_idx)
                    if link_value:
                        product_name_cell.hyperlink = str(link_value)

                target_row_idx += 1

def main():
    workbook_path = "Nirvana.xlsx"
    wb = openpyxl.load_workbook(workbook_path)

    # 1) "Sheet1" -> "Genel Rapor"
    source_sheet = wb["Sheet1"]
    source_sheet.title = "Genel Rapor"

    # 2) İki kopya oluştur
    rpt_sheet = wb.copy_worksheet(source_sheet)
    rpt_sheet.title = "RPT Raporu"

    indirim_sheet = wb.copy_worksheet(source_sheet)
    indirim_sheet.title = "İndirim Raporu"

    # RPT Raporu sayfasında en başa boş bir sütun ekleyip A1 hücresine "Sipariş Adedi" yazalım.
    rpt_sheet.insert_cols(1)
    rpt_sheet["A1"] = "Sipariş Adedi"

    # "Sipariş Adedi" kolonunun genişliğini 103px'e ayarlayalım.
    # Excel sütun genişlik birimi yaklaşık (px - 5) / 7 formülü ile hesaplanır.
    # (103 - 5) / 7 ≈ 14
    siparis_width = 14
    rpt_sheet.column_dimensions["A"].width = siparis_width

    # A sütunundaki tüm hücreleri, hem yatay hem dikey olarak ortalayalım.
    for row in range(1, rpt_sheet.max_row + 1):
        cell = rpt_sheet.cell(row=row, column=1)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # "Ürün Adı" kolonunun genişliğini 467px'e ayarlayalım.
    # (467 - 5) / 7 ≈ 66
    target_width = 66

    # İlk satırdaki hücreler arasında "Ürün Adı" başlığını arıyoruz.
    for cell in rpt_sheet[1]:
        if cell.value == "Ürün Adı":
            col_letter = cell.column_letter
            rpt_sheet.column_dimensions[col_letter].width = target_width
            break

    # 3) Kolon gizlemeleri
    # a) Genel Rapor
    hide_columns_by_header(
        wb["Genel Rapor"],
        ["Resim", "Marka", "Asorti", "StokKodu", "Üründe Hareket Var mı?"]
    )

    # b) RPT Raporu
    rpt_hide_cols = [
        "Stok Adedi Site ve Vega",
        "Ortalama Görüntülenme Adedi",
        "Kaç Güne Biter Site ve Vega",
        "GMT Stok Adedi",
        "SİTA Stok Adedi",
        "Mevsim",
        "Son Transfer Tarihi",
        "Son İndirim Tarihi",
        "Resim",
        "Marka",
        "Liste Fiyatı",
        "Asorti",
        "StokKodu",
        "Üründe Hareket Var mı?",
    ]
    hide_columns_by_header(wb["RPT Raporu"], rpt_hide_cols)

    # c) İndirim Raporu (ilk etapta 'Resim' ve 'Marka' kolonlarını gizle)
    hide_columns_by_header(
        wb["İndirim Raporu"],
        ["Resim", "Marka", "Üründe Hareket Var mı?", "Stok Adedi Site ve Vega", "Kaç Güne Biter Site ve Vega", "Liste Fiyatı", "GMT Stok Adedi", "SİTA Stok Adedi", "Mevsim", "Net Satış Tarihi ve Adedi"," Asorti", "StokKodu"]
    )

    # 4) "İndirim Raporu" sayfasında filtreleme
    temp_sheet_name = "Temp_Indirim"
    if temp_sheet_name in wb.sheetnames:
        del wb[temp_sheet_name]

    temp_sheet = wb.create_sheet(temp_sheet_name)

    filter_indirim_sheet_with_styles(
        source_sheet=wb["İndirim Raporu"],
        target_sheet=temp_sheet,
        instagram_stok_header="Instagram Stok Adedi",
        resim_header="Resim",
        urun_adi_header="Ürün Adı"
    )

    # Orijinal "İndirim Raporu" sayfasını sil, temp'i yeniden adlandır
    del wb["İndirim Raporu"]
    temp_sheet.title = "İndirim Raporu"

    # Tekrar 'Resim' ve 'Marka' kolonlarını gizle (yeni sayfa)
    hide_columns_by_header(
        wb["İndirim Raporu"],
        ["Resim", "Marka", "Üründe Hareket Var mı?", "Stok Adedi Site ve Vega", "Kaç Güne Biter Site ve Vega", "Liste Fiyatı", "GMT Stok Adedi", "SİTA Stok Adedi", "Mevsim", "Net Satış Tarihi ve Adedi", "Asorti", "StokKodu"]
    )

    # 6) Başlık satırlarını dondur ve zoom ayarını %90 yap
    for sheet_name in ["Genel Rapor", "RPT Raporu", "İndirim Raporu"]:
        sh = wb[sheet_name]
        freeze_header(sh)
        sh.sheet_view.zoomScale = 90

    # 7) RPT Raporu ve İndirim Raporu sayfalarına tablo stili ekle
    apply_table_format(wb["RPT Raporu"], table_name="RPTTable", style_name="TableStyleMedium9")
    apply_table_format(wb["İndirim Raporu"], table_name="IndirimTable", style_name="TableStyleMedium9")

    # 8) Kaydet
    wb.save(workbook_path)

    # 9) 3 sayfada da 'Kar Yüzdesi' kolonundaki verileri 100 ile çarp
    for sheet_name in ["Genel Rapor", "RPT Raporu", "İndirim Raporu"]:
        sheet = wb[sheet_name]
        # Önce "Kar Yüzdesi" kolonunu bul
        kar_col_idx = None
        for col_index in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=col_index).value == "Kar Yüzdesi":
                kar_col_idx = col_index
                break

        if kar_col_idx:
            # 2. satırdan son satıra kadar değeri 100 ile çarp
            for row_index in range(2, sheet.max_row + 1):
                val = sheet.cell(row=row_index, column=kar_col_idx).value
                if isinstance(val, (int, float)):
                    sheet.cell(row=row_index, column=kar_col_idx).value = val * 100

    # -------------------------------------------------------------------------
    # İSTENEN EKLEME: 3 sayfada da "Liste Fiyatı" ve "Satış Fiyatı" farkı %5'i aşıyorsa
    # "Alış Fiyatı" hücresinin yazı rengini kırmızı yapalım
    # -------------------------------------------------------------------------
    for sheet_name in ["Genel Rapor", "RPT Raporu", "İndirim Raporu"]:
        sheet = wb[sheet_name]

        # Kolon indekslerini bulalım
        liste_fiyati_idx = None
        satis_fiyati_idx = None
        alis_fiyati_idx = None

        for col_index in range(1, sheet.max_column + 1):
            header_val = sheet.cell(row=1, column=col_index).value
            if header_val == "Liste Fiyatı":
                liste_fiyati_idx = col_index
            elif header_val == "Satış Fiyatı":
                satis_fiyati_idx = col_index
            elif header_val == "Alış Fiyatı":
                alis_fiyati_idx = col_index

        # Eğer bu kolonlar varsa kontrol edelim
        if liste_fiyati_idx and satis_fiyati_idx and alis_fiyati_idx:
            for row_index in range(2, sheet.max_row + 1):
                try:
                    list_val = float(sheet.cell(row=row_index, column=liste_fiyati_idx).value or 0)
                    satis_val = float(sheet.cell(row=row_index, column=satis_fiyati_idx).value or 0)
                    
                    # Liste Fiyatı 0 değilse hesaplama yapalım
                    if list_val != 0:
                        fark = list_val - satis_val
                        yuzde = (fark / list_val) * 100
                        if yuzde > 5:
                            # Alış Fiyatı hücresinin yazı rengini kırmızıya çevir
                            cell_alis = sheet.cell(row=row_index, column=alis_fiyati_idx)
                            existing_font = copy(cell_alis.font)
                            existing_font.color = "FF0000"  # Kırmızı
                            cell_alis.font = existing_font
                except:
                    # Herhangi bir hatada (None, string vs.), atla
                    pass

    # Tüm değişiklikleri kaydet
    wb.save(workbook_path)

if __name__ == "__main__":
    main()










# Dosyayı yükle
workbook = openpyxl.load_workbook("Nirvana.xlsx")
worksheet = workbook["RPT Raporu"]

# Başlık satırındaki kolon isimlerini ve indekslerini tespit edelim (1. satır)
kolon_indexleri = {}
for idx, hucre in enumerate(worksheet[1], start=1):
    kolon_indexleri[hucre.value] = idx

# "Liste Fiyatı" ve "Satış Fiyatı" kolonlarının indekslerini alalım
liste_fiyati_col = kolon_indexleri.get("Liste Fiyatı")
satis_fiyati_col = kolon_indexleri.get("Satış Fiyatı")

if liste_fiyati_col is None or satis_fiyati_col is None:
    raise ValueError("Gerekli kolonlardan biri bulunamadı: 'Liste Fiyatı' veya 'Satış Fiyatı'.")

# Yeni kolonun (İndirim Oranı) ekleneceği indeks
yeni_kolon = worksheet.max_column + 1
worksheet.cell(row=1, column=yeni_kolon, value="İndirim Oranı")

# Veriler üzerinde işlem yapalım (başlık satırından sonraki satırlar)
for satir in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
    liste_deger = satir[liste_fiyati_col - 1].value  # 0 tabanlı indeksleme
    satis_deger = satir[satis_fiyati_col - 1].value

    # Liste Fiyatı değerinin geçerli ve sıfırdan farklı olduğundan emin olalım
    if liste_deger is not None and satis_deger is not None and liste_deger != 0:
        indirim_orani = (liste_deger - satis_deger) / liste_deger * 100
        worksheet.cell(row=satir[0].row, column=yeni_kolon, value=indirim_orani)
        
        # Eğer indirim oranı 10 veya daha büyükse satırı gizle
        if indirim_orani >= 10:
            worksheet.row_dimensions[satir[0].row].hidden = True
    else:
        # Geçerli veri yoksa yeni kolonda boş bırak
        worksheet.cell(row=satir[0].row, column=yeni_kolon, value=None)

# Yeni oluşturduğumuz "İndirim Oranı" kolonunu gizleyelim
col_letter = get_column_letter(yeni_kolon)
worksheet.column_dimensions[col_letter].hidden = True

# Sonucu yeni bir dosyaya kaydedelim
workbook.save("Nirvana.xlsx")











# Açıklama metinlerini içeren sözlük
aciklamalar = {
    "Ürün Adı": "Ürünün satıştaki adını belirtir.",
    "Üründe Hareket Var mı?": 'Bu alanda "Hayır" seçeneğin işareti kaldırılırsa, üründen hiçbir yerde yok ise ve aynı zamanda satışında da yakın zamanda bir hareket yok ise bu satırları gizle anlamına gelecektir.',
    "Instagram Stok Adedi": "Satıştaki stok adedini belirtir.",
    "Stok Adedi Her Şey Dahil": "Satıştaki stok, depodaki stok, çuvaldaki stok, siparişteki stokların toplamını belirtir.",
    "Stok Adedi Site ve Vega": "Satıştaki stokla depoda satışa hazır ürünlerin stoğunu belirtir.",
    "Günlük Ortalama Satış Adedi": "Son 7 günde ürünün satışta kaldığı güne göre ortalama bir satış adedi belirtir.",
    "Dünün Satış Adedi": "Ürünün dün kaç adet sattığını belirtir.",
    "Ortalama Görüntülenme Adedi": "Son 7 günde ürünün satışta kaldığı güne göre ortalama bir görüntülenme adedi belirtir.",
    "Görüntülenmenin Satışa Dönüş Oranı": "Son 7 günde ürünün satışta kaldığı güne göre ortalama bir görüntülenme adedi belirlenir ve bu görüntülenme adedine göre yine son 7 gündeki ortalama satışı baz alınarak bir yüzde hesabı çıkarılır.",
    "Kaç Güne Biter Her Şey Dahil": "Her dahil stokların, ortalama satış adedi baz alınarak kaç güne biteceğin belirtir.",
    "Kaç Güne Biter Site ve Vega": "Sadece satıştaki ve depoda satışa hazır stokların ortalama satış adedi baz alınarak kaç güne biteceğini belirtir.",
    "Alış Fiyatı": "Ürünün siteye geçirilen alış fiyatını belirtir. NOT : Ürünü alırkenki alış fiyatını yansıtmayabilir.",
    "Son Satın Alma Fiyatı": "Ürünün sipariş verirken yazılan son alış fiyatını belirtir.",
    "Satış Fiyatı": "Ürünün aktif satış fiyatını belirtir.",
    "Liste Fiyatı": "Ürünün alış fiyatına oranla olması gereken fiyatını belirtir.",
    "Kar Yüzdesi": "Ürünün yüzde kaç karla sattığını belirtir ve hesaplaması şu şekildedir (Satış - Alış) / Satış) * 100",
    "Resim Yüklenme Tarihi": "Ürünün resminin yüklendiği bir nevi satışa açıldığı tarihi belirtir.",
    "Kategori": "Ürünün ana kategorisini belirtir.",
    "GMT Stok Adedi": "Üründen çuvalda kaç adet olduğunu belirtir.",
    "SİTA Stok Adedi": "Ürünün siparişte kaç adet olduğunu belirtir.",
    "Mevsim": "Ürünün mevsimini belirtir.",
    "Net Satış Tarihi ve Adedi": "Ürünün tüm renk ve bedenlerinin aynı anda satışta olduğu son günün satış tarihi ve adedini belirtir.",
    "Son Transfer Tarihi": "Ürünün Instagram depoya en son ne zaman transfer edildiğini belirtir.",
    "Son İndirim Tarihi": "Ürüne en son ne zaman indirim yapıldığını belirtir.",
    "Etiketler": "Ürüne tanımlanan etiketleri gösterir mesela paça tipi gibi",
    "Bedenler": "Ürünün bedenlerinin ve anlık olarak o bedenlerden kaçar adet kaldığını belirtir"
}

# Dosyayı aç
workbook = openpyxl.load_workbook("Nirvana.xlsx")

# Tüm sayfaları tarayalım
for worksheet in workbook.worksheets:
    # İlk satırdaki hücrelerde başlıklar olduğunu varsayıyoruz
    for cell in worksheet[1]:
        if cell.value in aciklamalar:
            # Yorum objesini oluşturup genişlik ve yükseklik ayarını yapalım
            comment = Comment(text=aciklamalar[cell.value], author="Auto")
            comment.width = 200   # 200 piksel genişlik
            comment.height = 200  # 200 piksel yükseklik
            cell.comment = comment

# Sonuç dosyasını kaydedelim
workbook.save("Nirvana.xlsx")














def add_hyperlinks_preserve_style():
    workbook_path = "Nirvana.xlsx"
    wb = openpyxl.load_workbook(workbook_path)
    
    # "RPT Raporu" sayfasının varlığını kontrol ediyoruz.
    if "RPT Raporu" not in wb.sheetnames:
        print("RPT Raporu sayfası bulunamadı.")
        return
    sheet = wb["RPT Raporu"]
    
    # İlk satırdaki hücrelerden "Resim" ve "Ürün Adı" kolonlarının indekslerini buluyoruz.
    resim_col = None
    urun_adi_col = None
    for cell in sheet[1]:
        if cell.value == "Resim":
            resim_col = cell.column
        elif cell.value == "Ürün Adı":
            urun_adi_col = cell.column
    
    if resim_col is None or urun_adi_col is None:
        print("Gerekli kolonlardan biri bulunamadı ('Resim' veya 'Ürün Adı').")
        return
    
    # 2. satırdan itibaren her satırda, "Resim" kolonundaki URL'yi "Ürün Adı" hücresine hyperlink olarak ekleyelim.
    for row in range(2, sheet.max_row + 1):
        resim_value = sheet.cell(row=row, column=resim_col).value
        if resim_value and isinstance(resim_value, str) and resim_value.strip():
            urun_adi_cell = sheet.cell(row=row, column=urun_adi_col)
            urun_adi_cell.hyperlink = resim_value.strip()
            # Mevcut hücre stilini değiştirmeden hyperlink ekliyoruz.
    
    wb.save(workbook_path)
    wb.close()

if __name__ == "__main__":
    add_hyperlinks_preserve_style()









# 1. Google Sheets'teki "Stok Faaliyet Raporu" sayfasını Excel formatında indirme (fiziksel dosya oluşturulmuyor)
google_sheet_id = "1UCfKTxoleZCCBGsdEiDy7Kk2fR1VUOeO3LnBkD7oNMA"
sheet_gid = "886202732"  # "Stok Faaliyet Raporu" sayfasının gid değeri
export_url = f"https://docs.google.com/spreadsheets/d/{google_sheet_id}/export?format=xlsx&gid={sheet_gid}"

response = requests.get(export_url)
if response.status_code == 200:
    excel_bytes = response.content
    excel_file_like = BytesIO(excel_bytes)
else:
    raise Exception("Google Sheets sayfası indirilemedi, lütfen bağlantıyı kontrol ediniz.")

# 2. İndirilen Excel dosyasını DataFrame olarak oku
df = pd.read_excel(excel_file_like)

# 3. A kolonuna (ilk sütun) göre büyükten küçüğe sıralama
col_A = df.columns[0]
df_sorted = df.sort_values(by=col_A, ascending=False)

# 4. B kolonundaki verileri düzenleme:
#    Her değeri string'e çevirip, noktalardan bölüyoruz; ilk iki parçayı alıp "m1." ön ekini ekliyoruz.
col_B = df.columns[1]

def process_value(val):
    s = str(val)
    parts = s.split('.')
    if len(parts) >= 2:
        return "m1." + ".".join(parts[:2])
    else:
        return "m1." + s

df_sorted[col_B] = df_sorted[col_B].apply(process_value)

# 5. İşlenmiş verileri Nirvana.xlsx dosyasına, yeni oluşturulan "ProcessedData" adlı sayfaya yazma ve sayfayı gizleme
excel_file = "Nirvana.xlsx"
new_sheet_name = "ProcessedData"

if not os.path.exists(excel_file):
    # Dosya mevcut değilse yeni dosya oluşturup yazıyoruz
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
         df_sorted.to_excel(writer, sheet_name=new_sheet_name, index=False)
         # Yeni eklenen sayfayı gizleme
         worksheet = writer.book[new_sheet_name]
         worksheet.sheet_state = 'hidden'
else:
    # Dosya mevcutsa, yeni sayfa ekleyip gizli hale getiriyoruz
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
         df_sorted.to_excel(writer, sheet_name=new_sheet_name, index=False)
         worksheet = writer.book[new_sheet_name]
         worksheet.sheet_state = 'hidden'









def olustur_siparis_sayfasi(dosya_yolu="Nirvana.xlsx"):
    wb = load_workbook(dosya_yolu, data_only=False)

    # "Sipariş Sayfası" varsa sil
    if "Sipariş Sayfası" in wb.sheetnames:
        del wb["Sipariş Sayfası"]

    # Yeni sayfa oluştur
    ws = wb.create_sheet("Sipariş Sayfası")

    # Başlıklar (artık 15 kolon)
    headers = [
        "Tespit",               # A
        "Sipariş Adedi Tespit", # B
        "Stok Kodu Tespit",     # C
        "Stok Kodu",            # D
        "Tam Ürün Adı",         # E
        "Firma Kodu",           # F
        "Ürün Kodu",            # G
        "Ürün Adı",             # H
        "Renk",                 # I
        "Resim",                # J
        "Sipariş Adedi",        # K
        "Fiyat",                # L
        "Asorti",               # M
        "Firmaya Not",

        "Kategori",             # N
        "Dönem"                 # O
    ]

    # 1. satır: Başlıkları ekle ve biçimlendir
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ###################################################################
    # 1) İlk 3 kolonu (A, B, C) 2'den 25.000'e kadar formüllerle doldur
    ###################################################################

    # A (Tespit) => =IF(B2="","",ROW(A1))
    for row in range(2, 25001):
        formula = f'=IF(B{row}="", "", ROW(A{row-1}))'
        ws.cell(row=row, column=1).value = formula

    # B (Sipariş Adedi Tespit) => ='RPT Raporu'!A2 vb.
    for row in range(2, 25001):
        formula = f'=IF(\'RPT Raporu\'!A{row}="", "", \'RPT Raporu\'!A{row})'
        ws.cell(row=row, column=2).value = formula

    # C (Stok Kodu Tespit) => ='RPT Raporu'!B2 vb.
    for row in range(2, 25001):
        formula = f"='RPT Raporu'!B{row}"
        ws.cell(row=row, column=3).value = formula

    ###################################################################
    # Sonraki kolonlar (D - O) 2'den 500'e kadar formüllerle doldur
    ###################################################################

    # D (Stok Kodu)
    for row in range(2, 501):
        formula = f'=IFERROR(VLOOKUP(SMALL(A:A, ROW(A{row-1})), A:C, 3, 0), "")'
        ws.cell(row=row, column=4).value = formula

    # E (Tam Ürün Adı)
    for row in range(2, 501):
        formula = f"=IFERROR(VLOOKUP(D{row}, 'RPT Raporu'!B:C, 2, 0), \"\")"
        ws.cell(row=row, column=5).value = formula

    # F (Firma Kodu)
    for row in range(2, 501):
        formula = f'=SUBSTITUTE(IFERROR(MID(E{row}, FIND(".", E{row}), 50), ""), ".", "")'
        ws.cell(row=row, column=6).value = formula

    # G (Ürün Kodu)
    for row in range(2, 501):
        formula = (
            f'=IFERROR('
            f'SUBSTITUTE('
            f'SUBSTITUTE('
            f'SUBSTITUTE(MID(E{row}, FIND(" - ", E{row}), 50)," - ",""),'
            f'F{row},""),'
            f'".",""),'
            f'""'
            f')'
        )
        ws.cell(row=row, column=7).value = formula

    # H (Ürün Adı)
    for row in range(2, 501):
        formula = (
            f'=IFERROR('
            f'LEFT('
            f'SUBSTITUTE(E{row}, " ", "xxq", LEN(E{row})-LEN(SUBSTITUTE(E{row}, " ", ""))-2),'
            f'FIND("xxq", SUBSTITUTE(E{row}, " ", "xxq", LEN(E{row})-LEN(SUBSTITUTE(E{row}, " ", ""))-2))-1'
            f'),'
            f'""'
            f')'
        )
        ws.cell(row=row, column=8).value = formula

    # I (Renk)
    for row in range(2, 501):
        formula = (
            f'=IFERROR('
            f'TRIM(LEFT(SUBSTITUTE(E{row}, H{row}, ""), '
            f'FIND(" - ", SUBSTITUTE(E{row}, H{row}, ""))-1)),' 
            f'""'
            f')'
        )
        ws.cell(row=row, column=9).value = formula

    # J (Resim)
    for row in range(2, 501):
        formula = f"=IFERROR(VLOOKUP(D{row}, 'RPT Raporu'!B:P, 15, 0), \"\")"
        ws.cell(row=row, column=10).value = formula

    # K (Sipariş Adedi)
    for row in range(2, 501):
        formula = f'=IFERROR(VLOOKUP(SMALL(A:A, ROW(A{row-1})), A:B, 2, 0), "")'
        ws.cell(row=row, column=11).value = formula

    # L (Fiyat)
    for row in range(2, 501):
        formula = f'=IF(D{row}="", "", VLOOKUP(D{row}, ProcessedData!B:C, 2, 0))'
        ws.cell(row=row, column=12).value = formula


    # L (Firmaya Not)
    for row in range(2, 501):
        ws.cell(row=row, column=14).value = "-"






    # M (Asorti)
    for row in range(2, 501):
        formula = f"=IFERROR(VLOOKUP(D{row}, 'RPT Raporu'!B:AE, 30, 0), \"\")"
        ws.cell(row=row, column=13).value = formula

    # N (Kategori)
    for row in range(2, 501):
        formula = f"=IFERROR(VLOOKUP(D{row}, 'RPT Raporu'!B:W, 22, 0), \"\")"
        ws.cell(row=row, column=15).value = formula

    # O (Dönem)
    for row in range(2, 501):
        formula = f"=PROPER(IFERROR(VLOOKUP(D{row}, 'RPT Raporu'!B:Z, 25, 0), \"\"))"
        ws.cell(row=row, column=16).value = formula

    # A, B, C, D ve E sütunlarını gizle
    ws.column_dimensions['A'].hidden = True
    ws.column_dimensions['B'].hidden = True
    ws.column_dimensions['C'].hidden = True
    ws.column_dimensions['D'].hidden = True
    ws.column_dimensions['E'].hidden = True

    # İstenen sütun genişliklerini ayarla
    ws.column_dimensions['F'].width = 15  # Firma Kodu
    ws.column_dimensions['G'].width = 15  # Ürün Kodu
    ws.column_dimensions['H'].width = 35  # Ürün Adı
    ws.column_dimensions['I'].width = 15  # Renk
    ws.column_dimensions['J'].width = 20  # Resim
    ws.column_dimensions['K'].width = 17  # Sipariş Adedi
    ws.column_dimensions['L'].width = 10  # Fiyat
    ws.column_dimensions['M'].width = 30  # Asorti
    ws.column_dimensions['N'].width = 15  # Kategori
    ws.column_dimensions['O'].width = 30  # Dönem
    ws.column_dimensions['P'].width = 15  # Dönem

    # Sayfa ölçeğini %85'e ayarla
    ws.sheet_view.zoomScale = 90

    # Filtre aralığı
    ws.auto_filter.ref = "A1:O25000"

    # Kaydet
    wb.save(dosya_yolu)

# Fonksiyonu çağırmak için:
olustur_siparis_sayfasi("Nirvana.xlsx")



























# -----------------------------------------------------------------------
# 1) Token alma ve multi-sayfa vendor verisi çekme
# -----------------------------------------------------------------------
def login():
    conn = http.client.HTTPSConnection("task.haydigiy.com")

    login_payload = {
        "apiKey": "MypGcaEInEOTzuYQydgDHQ",
        "secretKey": "jRqliBLDPke76YhL_WL5qg",
        "emailOrPhone": "mustafa_kod@haydigiy.com",
        "password": "123456"
    }
    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json"
    }

    conn.request("POST", "/api/customer/login", body=json.dumps(login_payload), headers=headers)
    res = conn.getresponse()
    data = res.read().decode("utf-8")
    if res.status != 200:
        print("Giriş başarısız:", data)
        return None

    login_data = json.loads(data)
    token = login_data.get("data", {}).get("token")
    if not token:
        print("Token alınamadı. Dönen veri:", login_data)
        return None
    return token

def get_all_vendor_data(token):
    """
    Tek seferde 100 kayıt geliyorsa, veri bitene kadar pageIndex arttırarak
    tüm vendor verisini toplar.
    """
    all_items = []
    page_index = 1
    while True:
        conn = http.client.HTTPSConnection("task.haydigiy.com")

        list_payload = {
            "manufacturerName": "",
            "published": None,
            "pageIndex": page_index,
            "pageSize": 100
        }
        list_headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
            "Accept": "application/json"
        }

        conn.request("POST", "/adminapi/vendor/list", body=json.dumps(list_payload), headers=list_headers)
        res = conn.getresponse()
        data = res.read().decode("utf-8")
        if res.status != 200:
            print(f"Sayfa {page_index} isteği başarısız:", data)
            break

        response_data = json.loads(data)
        items = response_data.get("data", [])
        if not items:
            # Boş geldi, bitir
            break

        all_items.extend(items)

        # 100'den az geldiyse son sayfa kabul edilir
        if len(items) < 100:
            break

        page_index += 1


    return all_items

# -----------------------------------------------------------------------
# 2) Regex yardımı: Ürün Adı ve vendor "name" alanlarından firma kodu çıkarma
# -----------------------------------------------------------------------
def parse_firma_code(text):
    """
    Esnek regex:  \.(\d+)\.?$
    Örnek: 
      "Dabıl Kumaş - 6003.1247." -> 1247
      "6003.1247" -> 1247
    """
    if not isinstance(text, str):
        return None
    text = text.strip()
    match = re.search(r'\.(\d+)\.?$', text)
    if match:
        return match.group(1)
    return None

def parse_vendor_code(text):
    """
    Vendor "name" örnek: ".1806." -> 1806
    """
    return parse_firma_code(text)

# -----------------------------------------------------------------------
# 3) Kod eşleştirme: DataFrame tarafında (FirmaKodu -> VendorNotu)
# -----------------------------------------------------------------------
def get_vendor_note_for_code(code, vendor_dict):
    if not code:
        return ""
    return " // ".join(vendor_dict[code]) if code in vendor_dict else ""

# -----------------------------------------------------------------------
# Ana işlem
# -----------------------------------------------------------------------
def main():
    # 1) Login olup tüm vendor verisini çek
    token = login()
    if not token:
        return

    vendor_data_list = get_all_vendor_data(token)

    # 2) Vendor verilerini code->list haritası (vendor_dict) yap
    vendor_dict = {}
    for item in vendor_data_list:
        name_text = item.get("name", "")  # "name" alanında firma bilgisi
        code = parse_vendor_code(name_text)
        if code:
            vendor_dict.setdefault(code, []).append(name_text)

    # 3) Nirvana.xlsx içinde 3 çalışma sayfasını oku:
    sheet_names = ["Genel Rapor", "RPT Raporu", "İndirim Raporu"]
    df_sheets = pd.read_excel("Nirvana.xlsx", sheet_name=sheet_names)

    # df_sheets => { "Genel Rapor": DataFrame, "RPT Raporu": DataFrame, ... }

    # 4) Her sayfada Ürün Adı kolonundan kodu bul -> VendorNotu ekle
    for s_name in sheet_names:
        df = df_sheets[s_name]
        if "Ürün Adı" not in df.columns:
            print(f"Uyarı: {s_name} sayfasında 'Ürün Adı' kolonu bulunamadı!")
            continue

        df["FirmaKodu"] = df["Ürün Adı"].apply(parse_firma_code)
        df["VendorNotu"] = df["FirmaKodu"].apply(lambda c: get_vendor_note_for_code(c, vendor_dict))
        df_sheets[s_name] = df  # Geri koymaya gerek yok ama okuma kolaylığı için

    # 5) openpyxl ile Nirvana.xlsx'i aç, her sayfada "Ürün Adı" hücresine comment ekle
    wb = load_workbook("Nirvana.xlsx")

    for s_name in sheet_names:
        if s_name not in wb.sheetnames:
            print(f"Uyarı: {s_name} adlı bir sayfa Excel'de yok.")
            continue

        ws = wb[s_name]
        df = df_sheets[s_name]

        # Sütun başlık satırının 1. satırda olduğunu varsayıyoruz
        # "Ürün Adı" kolonunun Excel'deki sütun index'ini bulmak için:
        header_row = 1
        col_urun_adi = None
        max_col = ws.max_column
        for col_idx in range(1, max_col + 1):
            cell_value = ws.cell(row=header_row, column=col_idx).value
            if cell_value == "Ürün Adı":
                col_urun_adi = col_idx
                break

        if not col_urun_adi:
            print(f"Uyarı: {s_name} sayfasında 'Ürün Adı' başlığı bulunamadı.")
            continue

        # DataFrame'de 0-bazlı index, Excel'de 2. satırdan veri başlıyor
        for df_index, row_data in df.iterrows():
            vendor_note = row_data.get("VendorNotu", "")
            if not vendor_note.strip():
                continue  # boşsa comment eklemeyelim

            excel_row = df_index + 2  # 1.satır başlık, 2.satır verinin başlangıcı
            cell = ws.cell(row=excel_row, column=col_urun_adi)
            # Comment'i ekle (varsa üzerine yazar)
            cell.comment = Comment(text=vendor_note, author="VendorInfo")

    # 6) Kaydet
    wb.save("Nirvana.xlsx")


if __name__ == "__main__":
    main()








wb = openpyxl.load_workbook("Nirvana.xlsx")

def px_to_width(px: int) -> float:
    return round(px / 7, 2)  # yaklaşık dönüşüm

# Hedef genişlikler (px)
extra_widths_px = {
    "Resim Yüklenme Tarihi": 150,
    "Etiketler": 363,
    "Özel Kategoriler": 706,
    "Sigara Ürün Ekle / Çıkar": 120,
    "Kategori": 216,
}

# -------------------------------------------------
# Genel Rapor sayfası
# -------------------------------------------------
ws = wb["Genel Rapor"]
# Bedenler (43)
ws.column_dimensions[[c.column_letter for c in ws[1] if c.value == "Bedenler"][0]].width = 43
# Diğer hedef genişlikler
for cell in ws[1]:
    if cell.value in extra_widths_px:
        ws.column_dimensions[cell.column_letter].width = px_to_width(extra_widths_px[cell.value])

# -------------------------------------------------
# RPT Raporu sayfası
# -------------------------------------------------
ws = wb["RPT Raporu"]
# Bedenler (43)
ws.column_dimensions[[c.column_letter for c in ws[1] if c.value == "Bedenler"][0]].width = 43
# Instagram Stok Adedi (15)
ws.column_dimensions[[c.column_letter for c in ws[1] if c.value == "Instagram Stok Adedi"][0]].width = 15
# Diğer hedef genişlikler
for cell in ws[1]:
    if cell.value in extra_widths_px:
        ws.column_dimensions[cell.column_letter].width = px_to_width(extra_widths_px[cell.value])

# -------------------------------------------------
# İndirim Raporu sayfası
# -------------------------------------------------
ws = wb["İndirim Raporu"]
# Bedenler (43)
ws.column_dimensions[[c.column_letter for c in ws[1] if c.value == "Bedenler"][0]].width = 43
# Diğer hedef genişlikler
for cell in ws[1]:
    if cell.value in extra_widths_px:
        ws.column_dimensions[cell.column_letter].width = px_to_width(extra_widths_px[cell.value])

wb.save("Nirvana.xlsx")






# ---------------- PATCH: Re‑calculate and ensure 'Liste Fiyatı' is filled ---------------- #
import openpyxl
from openpyxl.styles import numbers

def _recalculate_liste_fiyati(workbook_path: str = "Nirvana.xlsx") -> None:
    """
    Ensure the 'Liste Fiyatı' column contains the correct calculated values on all three
    report sheets. This patch is placed at the *very* end of the script so that no later
    operation can overwrite it.
    """
    wb = openpyxl.load_workbook(workbook_path)
    target_sheets = [s for s in ["Genel Rapor", "RPT Raporu", "İndirim Raporu"] if s in wb.sheetnames]

    def calc_price(alis: float, kategori: str | None) -> float:
        if 0 <= alis <= 24.99:
            result = alis + 10
        elif 25 <= alis <= 39.99:
            result = alis + 13
        elif 40 <= alis <= 59.99:
            result = alis + 17
        elif 60 <= alis <= 200.99:
            result = alis * 1.35
        else:
            result = alis * 1.30

        if isinstance(kategori, str) and any(k in kategori for k in ["Parfüm", "Gözlük", "Saat", "Kolye", "Küpe", "Bileklik", "Bilezik"]):
            result *= 1.20
        else:
            result *= 1.10

        return int(round(result)) + 0.99

    for ws in (wb[s] for s in target_sheets):
        headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        alis_idx   = headers.get("Alış Fiyatı") or headers.get("AlisFiyati")
        liste_idx  = headers.get("Liste Fiyatı") or headers.get("ListeFiyatı")
        kategori_idx = headers.get("Kategori")
        if not (alis_idx and liste_idx):
            continue

        for row in range(2, ws.max_row + 1):
            cell_alis = ws.cell(row=row, column=alis_idx).value
            if cell_alis is None:
                continue
            try:
                alis_val = float(cell_alis)
            except Exception:
                continue
            kategori_val = ws.cell(row=row, column=kategori_idx).value if kategori_idx else ""
            liste_val = calc_price(alis_val, kategori_val)
            cell_liste = ws.cell(row=row, column=liste_idx)
            cell_liste.value = liste_val
            cell_liste.number_format = '#,##0.00₺'

    wb.save(workbook_path)

if __name__ == "__main__":
    _recalculate_liste_fiyati()
# ---------------- END PATCH ------------------------------------------------------------- #
